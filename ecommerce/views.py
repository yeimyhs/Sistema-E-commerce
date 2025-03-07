from rest_framework.permissions import AllowAny
from rest_framework.viewsets import ModelViewSet
from ecommerce.serializers import  AdministracionSerializer, CuponSerializer, MarcaSerializer, MonedaSerializer, PromocionSerializer, TblcarritoSerializer, TblitemSerializer, TblnoticiaSerializer, TblpedidoSerializer, TblCarruselSerializer, TipocambioSerializer, ValoracionSerializer, TbldetallecarritoSerializer, TblimagenitemSerializer, TblitemclaseSerializer, TblitempropiedadSerializer, TblitemrelacionadoSerializer, TbldetallepedidoSerializer
from ecommerce.models import Administracion, Cupon, Marca, Moneda, Promocion, Tblcarrito, Tblitem, Tblnoticia, Tblpedido, TblCarrusel, Tipocambio, Valoracion, Tbldetallecarrito, Tblimagenitem, Tblitemclase, Tblitempropiedad, Tblitemrelacionado, Tbldetallepedido

from django.contrib.auth import login
from rest_framework import permissions
from rest_framework.authtoken.serializers import AuthTokenSerializer
from knox.views import LoginView as KnoxLoginView
from rest_framework import generics

from rest_framework import filters

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from knox.models import AuthToken

from .models import *
from .serializers import *
from .filters import *

from rest_framework.decorators import action
from rest_framework import filters
#from .filters import TblitemFilter

from django.db.models import Q
from rest_framework import viewsets
from rest_framework.response import Response

from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi


import rarfile
import os
from django.http import JsonResponse, HttpResponseBadRequest
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.views.decorators.csrf import csrf_exempt
from .models import Tblitem
import os
import rarfile
from io import BytesIO
from PIL import Image
from django.core.files import File
from django.http import JsonResponse
from rest_framework.decorators import api_view
from .models import Tblitem
from django.conf import settings


from rest_framework.generics import ListAPIView


from ecommerce.permissions import *

rarfile.UNRAR_TOOL = r"C:\Program Files\WinRAR\UnRAR.exe"
@csrf_exempt
def upload_images(request):
    """
    Servicio POST para subir un archivo RAR que contiene imágenes y asociarlas con los productos usando el SKU.
    """
    # Verificar que el archivo RAR esté presente en la solicitud
    if 'rar_file' not in request.FILES:
        return JsonResponse({'error': 'No se proporcionó un archivo RAR'}, status=400)

    rar_file = request.FILES['rar_file']

    # Guardar temporalmente el archivo RAR en el sistema
    rar_file_path = os.path.join(settings.MEDIA_ROOT, 'temp.rar')
    with open(rar_file_path, 'wb') as f:
        for chunk in rar_file.chunks():
            f.write(chunk)

    try:
        # Abrir el archivo RAR y procesar las imágenes
        with rarfile.RarFile(rar_file_path) as rf:
            # Iterar sobre los archivos dentro del RAR
            for archivo in rf.infolist():
                if archivo.is_file() and archivo.filename.lower().endswith('.jpg'):
                    # Obtener el SKU del nombre de la imagen
                    nombre_imagen = archivo.filename
                    sku_imagen = nombre_imagen.split('/')[-1].split('.')[0]  # Eliminar cualquier prefijo y la barra
 # El SKU es todo antes de .jpg
                    print(sku_imagen)
                    # Buscar el producto en la base de datos por el SKU
                    try:
                        producto = Tblitem.objects.get(codigosku=sku_imagen)
                    except Tblitem.DoesNotExist:
                        print(f"No se encontró el producto con SKU {sku_imagen}")
                        continue

                    # Extraer la imagen
                    with rf.open(archivo) as f:
                        imagen_bytes = f.read()
                        imagen = Image.open(BytesIO(imagen_bytes))

                        # Guardar la imagen en el campo imagenprincipal del producto
                        imagen.name = archivo.filename  # Asignar el nombre original de la imagen
                        imagen_path = os.path.join(settings.MEDIA_ROOT, 'imagenPrincipalItem', nombre_imagen)

                        # Guardar la imagen en el sistema de archivos
                        imagen.save(imagen_path)

                        # Asignar la imagen al campo 'imagenprincipal' del producto
                        with open(imagen_path, 'rb') as image_file:
                            producto.imagenprincipal = File(image_file, nombre_imagen)
                            producto.save()

                    print(f"Imagen para SKU {sku_imagen} guardada exitosamente.")
        
        # Eliminar el archivo temporal
        os.remove(rar_file_path)

        return JsonResponse({'message': 'Imágenes procesadas y asociadas correctamente'}, status=200)

    except Exception as e:
        # Manejo de errores si ocurre algún problema
        return JsonResponse({'error': str(e)}, status=500)

import requests
from requests.auth import HTTPBasicAuth
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json


@csrf_exempt
def create_payment(request):
    if request.method == 'POST':
        try:
            payload = json.loads(request.body)  # Asegúrate de convertir el cuerpo a JSON
            url = "https://api.micuentaweb.pe/api-payment/V4/Charge/CreatePayment"
            
            # Imprimir el payload para verificar que estás enviando los datos correctamente
            print("Payload:", payload)

            response = requests.post(
                url,
                auth=HTTPBasicAuth(settings.IZIPAY_USERNAME, settings.IZIPAY_PASSWORD),
                json=payload
            )
            
            # Imprimir la respuesta para ver qué devuelve el servidor
            print("Response Status Code:", response.status_code)
            print("Response Body:", response.text)

            if response.status_code == 200:
                return JsonResponse(response.json(), status=200)
            else:
                # Si hay error, devolver el mensaje y el código de estado
                return JsonResponse({
                    "status": "error", 
                    "message": response.json()
                }, status=400)
        
        except Exception as e:
            # Capturar el error completo y devolverlo en la respuesta
            print(f"Error: {e}")
            return JsonResponse({"status": "error", "message": "Hubo un error al procesar la solicitud."}, status=500)
    
    else:
        return JsonResponse({"status": "error", "message": "Método no permitido."}, status=405)
    
    
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import requests
from requests.auth import HTTPBasicAuth
import json

# URL del servicio externo

# Credenciales

@csrf_exempt
def generate_token(request):
    """
    Vista para generar un token a través del servicio CreateToken.
    "currency": "PEN",  # Ejemplo de moneda
    "customer": {
        "email":"ejemplo@otro.com"
    }
    """
    if request.method == "POST":
        try:
            # Parsear el cuerpo de la solicitud
            payload = json.loads(request.body)
            TOKEN_URL = "https://api.micuentaweb.pe/api-payment/V4/Charge/CreateToken"
            
            # Enviar solicitud POST al servicio externo
            response = requests.post(
                TOKEN_URL,
                auth=HTTPBasicAuth(settings.IZIPAY_USERNAME, settings.IZIPAY_PASSWORD),
                #auth=HTTPBasicAuth(USERNAME, PASSWORD),
                json=payload
            )
            
            # Verificar el estado de la respuesta
            if response.status_code == 200:
                return JsonResponse(response.json(), status=200)
            else:
                return JsonResponse({
                    "error": "Error al generar el token",
                    "status_code": response.status_code,
                    "details": response.text
                }, status=response.status_code)
        
        except Exception as e:
            # Manejar errores y retornar una respuesta apropiada
            return JsonResponse({"error": str(e)}, status=500)
    
    # Método no permitido
    return JsonResponse({"error": "Método no permitido"}, status=405)


@csrf_exempt
def create_token(request):
    """
    Vista para la creación de un token de tarjeta.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            url = f"{BASE_API_URL}tokens"
            response = requests.post(url, headers=HEADERS, json=data)
            return JsonResponse(response.json(), status=response.status_code)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Método no permitido"}, status=405)

@csrf_exempt
def get_token(request):
    """
    Vista para consultar datos de un token específico.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            url = f"{BASE_API_URL}tokens/token"
            response = requests.post(url, headers=HEADERS, json=data)
            return JsonResponse(response.json(), status=response.status_code)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Método no permitido"}, status=405)

@csrf_exempt
def get_tokens_by_buyer(request):
    """
    Vista para consultar todos los tokens registrados para un comprador.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            url = f"{BASE_API_URL}tokens/tokens"
            response = requests.post(url, headers=HEADERS, json=data)
            return JsonResponse(response.json(), status=response.status_code)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Método no permitido"}, status=405)




import hmac
import hashlib
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

import hashlib

from cryptography.hazmat.primitives import hmac, hashes
from cryptography.hazmat.backends import default_backend


# Define tu clave secreta
SHA_KEY = settings.IZIPAY__KEY  # Reemplaza con tu clave HMAC-SHA-256
def check_hash(data, key):
    """
    Verifica la firma del mensaje usando HMAC-SHA256
    """
    print(SHA_KEY)
    supported_sign_algos = ['sha256_hmac']
    alg = data.get('kr-hash-algorithm')

    # Validar el algoritmo de hash
    if alg not in supported_sign_algos:
        return False

    # Obtener el contenido de 'kr-answer' y asegurarse de que sea una cadena JSON
    kr_answer = json.dumps(data.get('kr-answer', {}), separators=(',', ':'))
    kr_answer = kr_answer.replace('\\/', '/')

    # Calcular el hash HMAC-SHA256
    calculated_hash = hmac.HMAC(key.encode('utf-8'),
                hashes.SHA256(),
                backend=default_backend())
    calculated_hash.update(kr_answer.encode('utf-8'))
    calculated_hash = calculated_hash.finalize().hex()

    # Comparar con el hash recibido
    received_hash = data.get('kr-hash')
    return calculated_hash == received_hash

from django.utils.timezone import now
from decimal import Decimal
from django.db import transaction


from django.core.mail import EmailMessage
from django.conf import settings
from django.utils.html import format_html


@csrf_exempt
@transaction.atomic
def process_payment(request):
    if request.method != 'POST':
        return HttpResponse('Método no permitido', status=405)
    
    pedido = None  # Definir la variable antes de usarla en cualquier excepción
    
    try:
        # Decodificar el JSON recibido
        data = json.loads(request.body.decode('utf-8'))
        
        # PASO 1: verificar la firma con SHA_KEY
        if not check_hash(data, SHA_KEY):
            return HttpResponse('Invalid signature.<br/>', status=400)

        # Preparar la respuesta siguiendo el formato original
        answer = {
            'message': 'OK',
            'kr-hash': data.get('kr-hash'),
            'kr-hash-algorithm': data.get('kr-hash-algorithm'),
            'kr-answer-type': data.get('kr-answer-type'),
            'kr-answer': data.get('kr-answer')
        }

        transaccion_data = data.get("kr-answer", {}).get("transactions", [])[0]
        card_details = transaccion_data.get("transactionDetails", {}).get("cardDetails", {})
        authorization_response = card_details.get("authorizationResponse", {})
        user_info = transaccion_data.get("transactionDetails", {}).get("userInfo", "DESCONOCIDO")
        idpedido = data.get("kr-answer", {}).get("orderDetails", {}).get("orderId")

        # Crear nueva transacción
        nueva_transaccion, created = tblTransaccion.objects.update_or_create(
            transaccion_id=transaccion_data.get("uuid"),
            defaults={
                "metodo_pago": transaccion_data.get("paymentMethodType", "DESCONOCIDO"),
                "nombre_en_tarjeta": card_details.get("cardHolderName") or user_info,
                "numero_tarjeta": card_details.get("pan", "XXXX XXXX XXXX XXXX"),
                "monto_total": Decimal(authorization_response.get("amount", 0)) / 100,
                "fecha_transaccion": authorization_response.get("authorizationDate", now()),
            }
        )
        try:
            with transaction.atomic():
            
                try:
                    pedido = Tblpedido.objects.get(idpedido=idpedido, activo=True)
                except Tblpedido.DoesNotExist:
                    return HttpResponse(json.dumps({"error": "El pedido no existe o no está activo."}), status=404, content_type="application/json")

                pedido.idtransaccion = nueva_transaccion
                pedido.estado = 2
                pedido.save()

                # Obtener los detalles del pedido
                detalles_pedido = Tbldetallepedido.objects.filter(idpedido=pedido)
                if not detalles_pedido.exists():
                    return HttpResponse(json.dumps({"error": "No hay detalles de pedido para procesar."}), status=404, content_type="application/json")

                # Procesar reducción de stock
                for detalle in detalles_pedido:
                        try:
                            item = detalle.idproduct
                        except Tblitem.DoesNotExist:
                            pedido.estado = 5  # Estado de error
                            pedido.save()
                            return HttpResponse(json.dumps({"error": f"Producto con ID {detalle.idproduct} no encontrado."}), status=404, content_type="application/json")

                        # Verificar si hay suficiente stock
                        if item.stock < detalle.cantidad:
                            pedido.estado = 5  # Estado de error
                            pedido.save()
                            return HttpResponse(json.dumps({
                                "error": f"Stock insuficiente para el producto {item.titulo} - SKU({item.codigosku} - {item.pk}). Disponible: {item.stock}, Requerido: {detalle.cantidad}"
                            }), status=400, content_type="application/json")

                        # Reducir stock
                        item.stock -= detalle.cantidad
                        item.save()
        except ValueError as e:
            pedido.estado = 5  # Estado de error
            pedido.save()
            return HttpResponse(json.dumps({"error": str(e)}), status=400, content_type="application/json")

        # Si todo salió bien, marcar como pagado (Estado 4)
        pedido.estado = 2
        pedido.save()




        #--------------------------------------------------------
        
        # Preparar los datos para el correo
        email_cliente = pedido.idcliente.email
        user_email  = email_cliente
        user_id  = pedido.idcliente.pk
        email_admin = settings.DEFAULT_FROM_EMAIL  # Correo del remitente configurado en settings
        total_pedido = nueva_transaccion.monto_total

        # Construir detalles del pedido en formato HTML
        detalles_html = "".join([
            f"<tr>"
            f"<td>{detalle.idproduct.titulo}</td>"
            f"<td>{detalle.cantidad}</td>"
            f"<td>${detalle.idproduct.precionormal:.2f}</td>"
            f"<td>${detalle.precioflete:.2f}</td>"
            f"</tr>"
            for detalle in detalles_pedido
        ])

        # Datos del pedido
        id_pedido = pedido.idpedido
        moneda = pedido.idmoneda.nombre if pedido.idmoneda else 'PEN'
        subtotal = pedido.subtotal
        # Calcular el costo de envío total sumando los precios de flete de cada detalle
        costo_envio = sum(detalle.precioflete for detalle in detalles_pedido if detalle.precioflete)

        descuento = pedido.totaldescuento
        total_pedido = nueva_transaccion.monto_total
        tipo_envio = dict(Tblpedido.TIPOS_ID_ENVIO).get(pedido.tipoenvio, 'No especificado')
        estado_pedido = pedido.estado

        # Datos de la transacción
        metodo_pago = nueva_transaccion.metodo_pago
        numero_tarjeta = nueva_transaccion.numero_tarjeta[-4:] if nueva_transaccion.numero_tarjeta else '****'

        # Datos de la sede
        sede_nombre = pedido.idsede.nombre if pedido.idsede else 'No especificado'
        sede_direccion = pedido.idsede.direccion if pedido.idsede else 'No disponible'

        # Construcción de los detalles del pedido en HTML
        detalles_html = "".join([
            f"<tr>"
            f"<td>{detalle.idproduct.titulo}</td>"
            f"<td>{detalle.cantidad}</td>"
            f"<td>${detalle.idproduct.precionormal:.2f}</td>"
            f"<td>${detalle.precioflete:.2f}</td>"
            f"</tr>"
            for detalle in detalles_pedido
        ])

# Cr

        # Crear cuerpo del correo en HTML
        html_mensaje_cliente = format_html(f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Confirmación de Pedido - Grupo IAP S.A.C</title>
</head>
<body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px; display: flex; justify-content: center;">
    <div style="max-width: 600px; width: 100%; background: white; border-radius: 8px; overflow: hidden; 
                box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); border-top: 6px solid rgb(254, 206, 0);">

        <!-- Encabezado -->
        <div style="padding: 25px; text-align: center; border-bottom: 2px solid #ddd;">
            <h2 style="color: #2c3e50; font-size: 24px; margin: 0;">🎉 ¡Gracias por tu compra en Grupo IAP S.A.C!</h2>
            <p style="color: #7a7a7a; font-size: 16px; margin-top: 5px;">Tu pedido ha sido recibido y está siendo procesado.</p>
        </div>

        <!-- Información del Pedido -->
        <div style="padding: 20px;">
            <h3 style="color: #2c3e50; font-size: 20px; border-bottom: 2px solid #ddd; padding-bottom: 5px;">Detalles del Pedido</h3>
            <p style="font-size: 16px;"><strong>Número de Orden:</strong> {id_pedido}</p>
            <p style="font-size: 16px;"><strong>Moneda:</strong> {moneda}</p>
            <p style="font-size: 16px;"><strong>Método de Pago:</strong> {metodo_pago} terminada en **{numero_tarjeta}</p>
        </div>

        <!-- Resumen del Pedido -->
        <div style="padding: 20px;">
            <h3 style="color: #2c3e50; font-size: 20px; border-bottom: 2px solid #ddd; padding-bottom: 5px;">Resumen del Pedido</h3>
            <table style="width: 100%; border-collapse: collapse; margin-top: 15px;">
                <tr style="background-color: rgb(254, 206, 0); color: black;">
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Producto</th>
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Cantidad</th>
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Precio Unitario</th>
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Costo de Envío</th>
                </tr>
                </thead>
                <tbody>
                    {detalles_html}
                </tbody>
            </table>

            <!-- Totales -->
            <!-- Totales -->
            <p style="margin-top: 15px; font-size: 18px; text-align: right;">
                <strong>Subtotal:</strong> {subtotal:,.2f} {moneda}<br>
                <strong>Costo de Envío:</strong> {costo_envio:,.2f} {moneda}<br>
                <strong>Descuento:</strong> -{descuento:,.2f} {moneda}<br>
                <strong style="font-size: 20px; color: black;">Total Pagado:</strong> <span style="color: black; font-size: 20px;">{total_pedido:,.2f} {moneda}</span>
            </p>

        </div>

        <!-- Información de Envío -->
        <div style="padding: 20px;">
            <h3 style="color: #2c3e50; font-size: 20px; border-bottom: 2px solid #ddd; padding-bottom: 5px;">Información de Envío</h3>
            <p style="font-size: 16px;"><strong>Dirección de Envío:</strong> {pedido.direcciondestino or 'No especificado'}</p>
            <p style="font-size: 16px;"><strong>Sede:</strong> {sede_nombre}</p>
            <p style="font-size: 16px;"><strong>Dirección de Sede:</strong> {sede_direccion}</p>
            <p style="font-size: 16px;"><strong>Tipo de Envío:</strong> {tipo_envio}</p>
            <p style="font-size: 16px;"><strong>Estado del Pedido:</strong> {estado_pedido}</p>
        </div>

        <!-- Contacto -->
        <div style="padding: 25px; text-align: center;">
            <p style="color: #555; font-size: 16px;">
                Si tienes alguna consulta, contáctanos en:<br>
                📧 <a href="mailto:soporte@grupoiap.com" style="color: #2c3e50; text-decoration: none; font-weight: bold;">soporte@grupoiap.com</a><br>
                📞 <strong>{settings.CONTACT_PHONE}</strong>
            </p>
            <p style="font-size: 18px; font-weight: bold;">Gracias por confiar en Grupo IAP S.A.C</p>
        </div>
    </div>
</body>
</html>
""")

        # Enviar correo al cliente si su email está disponible
        if email_cliente:
            email = EmailMessage(
                subject="Confirmación de compra",
                body=html_mensaje_cliente,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email_cliente],
            )
            email.content_subtype = "html"  # Importante para que el correo sea interpretado como HTML
            email.send()

        # Enviar correo al administrador con información del pedido
        html_mensaje_admin = format_html(f"""

<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Confirmación de Pedido - Grupo IAP S.A.C</title>
</head>
<body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px; display: flex; justify-content: center;">
    <div style="max-width: 600px; width: 100%; background: white; border-radius: 8px; overflow: hidden; 
                box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.1); border-top: 6px solid rgb(254, 206, 0);">

        <!-- Encabezado -->
        <div style="padding: 20px; text-align: center; border-bottom: 2px solid #ddd;">
            <h2 style="color: #2c3e50; margin: 0;">🛒 Nuevo Pedido Recibido</h2>
            <p style="color: #7a7a7a; margin-top: 5px;">Detalles del pedido a continuación:</p>
        </div>

        <!-- Sección de Información -->
        <div style="padding: 20px;">
            <div style="background-color: #f8f8f8; padding: 15px; border-radius: 5px;">
                <p style="margin: 5px 0; color: #333;"><strong>ID Pedido:</strong> <span style="color: rgb(254, 206, 0);">{pedido.idpedido}</span></p>
                <p style="margin: 5px 0; color: #333;"><strong>ID Usuario:</strong> {user_id}</p>
                <p style="margin: 5px 0; color: #333;"><strong>Correo Usuario:</strong> {user_email}</p>
                <p style="margin: 5px 0; color: #333;"><strong>Usuario:</strong> {user_info}</p>
                <p style="margin: 5px 0; color: #333;"><strong>Transacción:</strong> {nueva_transaccion.transaccion_id}</p>
            </div>


        <!-- Resumen del Pedido -->
        <div style="padding: 20px;">
            <h3 style="color: #2c3e50; font-size: 20px; border-bottom: 2px solid #ddd; padding-bottom: 5px;">Resumen del Pedido</h3>
            <table style="width: 100%; border-collapse: collapse; margin-top: 15px;">
                <tr style="background-color: rgb(254, 206, 0); color: black;">
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Producto</th>
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Cantidad</th>
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Precio Unitario</th>
                    <th style="border-bottom: 2px solid #ddd; padding: 8px;">Costo de Envío</th>
                </tr>
                </thead>
                <tbody>
                    {detalles_html}
                </tbody>
            </table>

            <!-- Totales -->
            <!-- Totales -->
            <p style="margin-top: 15px; font-size: 18px; text-align: right;">
                <strong>Subtotal:</strong> {subtotal:,.2f} {moneda}<br>
                <strong>Costo de Envío:</strong> {costo_envio:,.2f} {moneda}<br>
                <strong>Descuento:</strong> -{descuento:,.2f} {moneda}<br>
                <strong style="font-size: 20px; color: black;">Total Pagado:</strong> <span style="color: black; font-size: 20px;">{total_pedido:,.2f} {moneda}</span>
            </p>

        </div>

        <!-- Información de Envío -->
        <div style="padding: 20px;">
            <h3 style="color: #2c3e50; font-size: 20px; border-bottom: 2px solid #ddd; padding-bottom: 5px;">Información de Envío</h3>
            <p style="font-size: 16px;"><strong>Dirección de Envío:</strong> {pedido.direcciondestino or 'No especificado'}</p>
            <p style="font-size: 16px;"><strong>Sede:</strong> {sede_nombre}</p>
            <p style="font-size: 16px;"><strong>Dirección de Sede:</strong> {sede_direccion}</p>
            <p style="font-size: 16px;"><strong>Tipo de Envío:</strong> {tipo_envio}</p>
            <p style="font-size: 16px;"><strong>Estado del Pedido:</strong> {estado_pedido}</p>
        </div>

        <!-- Contacto -->
        <div style="padding: 25px; text-align: center;">
            <p style="color: #555; font-size: 16px;">
                Si tienes alguna consulta, contáctanos en:<br>
                📧 <a href="mailto:soporte@grupoiap.com" style="color: #2c3e50; text-decoration: none; font-weight: bold;">soporte@grupoiap.com</a><br>
                📞 <strong>{settings.CONTACT_PHONE}</strong>
            </p>
            <p style="font-size: 18px; font-weight: bold;">Gracias por confiar en Grupo IAP S.A.C</p>
        </div>
    </div>
</body>
</html>
""")



        email_admin = EmailMessage(
            subject=f"Nuevo pedido recibido  #{pedido.idpedido}",
            body=html_mensaje_admin,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_admin],
        )
        email_admin.content_subtype = "html"
        email_admin.send()
        #--------------------------------------
        return HttpResponse(json.dumps(answer), content_type='application/json')

    except json.JSONDecodeError:
        if pedido:  # Evitar el error si pedido aún no ha sido asignado
            pedido.estado = 5
            pedido.save()
        return HttpResponse(json.dumps({"error": "Formato JSON inválido."}), status=400, content_type="application/json")

    except Exception as e:
        if pedido:  # Evitar el error si pedido aún no ha sido asignado
            pedido.estado = 5
            pedido.save()
        return HttpResponse(json.dumps({"error": str(e)}), status=500, content_type="application/json")


from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Count
from django.db.models import Q, Exists, OuterRef


class ClasesYPropiedadesView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        # JSON de entrada
        filtros = request.data
        categoria_id = filtros.get('categoria')
  

        if not categoria_id:
            items_filtrados = Tblitem.objects.filter(activo=True
        )
            poblacion = items_filtrados
        else:
        # Obtener los items relacionados con la categoría
            items_filtrados = Tblitem.objects.filter(
                categoria_relacionada__idcategoria=categoria_id, activo=True
            )
            poblacion = items_filtrados
            

        if not items_filtrados.exists():
            return Response({'error': 'No hay ítems para esta categoría.'}, status=400)

        filtro_general = []
        clases_ids = tblitemclasevinculo.objects.filter(
            iditem__in=items_filtrados, activo=True
        ).values_list('idclase', flat=True).distinct()
        clases = Tblitemclase.objects.filter(idclase__in=clases_ids).prefetch_related('vinculos').distinct()
  
        for clase in clases:
            propiedades_list = tblitemclasevinculo.objects.filter(
                idclase=clase, activo=True
            ).values_list('propiedad', flat=True).distinct()  # Extraer solo nombres únicos de propiedades

            filtro_general.append({
                "id": clase.pk,
                "idclase": clase.idclase,
                "clase": clase.nombre,
                "propiedades": [{"nombre": p} for p in propiedades_list]  # Formato requerido
            })


        # Filtrar marcas en base a los items
        marcas_ids = items_filtrados.values_list('idmodelo__idmarca', flat=True).distinct()
        marcas = Marca.objects.filter(id__in=marcas_ids)
        lista_marcas = [{"id": marca.id, "nombre": marca.nombre} for marca in marcas]

        # Filtrar modelos en base a los items
        modelos_ids = items_filtrados.values_list('idmodelo', flat=True).distinct()
        modelos = Tblmodelo.objects.filter(id__in=modelos_ids)
        lista_modelos = [
            {"idmodelo": modelo.id, "nombre": modelo.nombre, "idmarca": modelo.idmarca.id if modelo.idmarca else None}
            for modelo in modelos
        ]

        # Filtrar valores de ancho en base a los items
        valores_ancho = items_filtrados.values_list('ancho', flat=True).distinct()
        valores_ancho = [ancho for ancho in valores_ancho if ancho is not None]

        # ADAPTACIÓN AL FORMATO ANTIGUO: 
        # En lugar de devolver marcas, modelos y anchos como estructuras separadas, 
        # las agregamos dentro de filtro_general como en la versión errónea.
        filtro_general.append({"marcas": lista_marcas})
        filtro_general.append({"modelos": lista_modelos})
        filtro_general.append({"anchos": valores_ancho}) 
        #-----------------------------------------------------------------
        #-----------------------------------------------------------------
        
        
        data = request.data
        ancho_filtro = data.get('ancho', None)
        perfil_filtro = data.get('perfil', None)
        filtro_dinamico = {
            "ancho": [],
            "perfil": [],
            "aro": []
        }

        # Obtener la población inicial
        #poblacion = Tblitem.objects.filter(activo=True)

        # **Nivel 1: Filtro por ancho**
        
        # Listado de valores distintos de ancho con su conteo
        anchos = poblacion.values('ancho').annotate(items_existentes=Count('idproduct')).order_by('ancho')
        filtro_dinamico["ancho"] = [
            {"ancho": ancho["ancho"], "items_existentes": ancho["items_existentes"]}
            for ancho in anchos
        ]
        if ancho_filtro:
            # Filtrar por ancho recibido
            #ancho_filtro = float(ancho_filtro)
            poblacion = poblacion.filter(ancho=ancho_filtro)
            # **Nivel 2: Filtro por perfil**
            # Obtener los vínculos relacionados con la clase "Perfil" (idclase=1)
            vinculados = tblitemclasevinculo.objects.filter(
                iditem__in=poblacion, 
                idclase__idclase=1, 
                activo=True
            )

            # Obtener perfiles únicos
            perfiles = vinculados.values('propiedad').annotate(items_existentes=Count('iditem', distinct=True))
            perfiles_data = [
                {"perfil": perfil["propiedad"] if perfil["propiedad"] != "-" else "(-)", 
                "items_existentes": perfil["items_existentes"]}
                for perfil in perfiles
            ]

            # Agregar los items sin vinculación o con propiedad "-"
            items_sin_vinculo = poblacion.exclude(
                clases_propiedades__idclase__idclase=1
            ).count()
            if items_sin_vinculo > 0:
                perfiles_data.append({"perfil": "(-)", "items_existentes": items_sin_vinculo})
            filtro_dinamico["perfil"] = perfiles_data

            # **Nivel 3: Filtro por aro**
            if perfil_filtro:
                if perfil_filtro == "(-)":
                    # Filtrar por items sin vinculación o con propiedad "-"
                    # Obtener ítems que tienen la propiedad "-" para la clase con idclase=1
                    items_con_propiedad_especial = poblacion.filter(
                        clases_propiedades__idclase__idclase=1,
                        clases_propiedades__propiedad="-"
                    ).values_list('idproduct', flat=True)

                    # Obtener ítems que no tienen ninguna relación con la clase con idclase=1
                    items_sin_relacion = poblacion.exclude(
                        clases_propiedades__idclase__idclase=1
                    ).values_list('idproduct', flat=True)

                    # Combinar ambos conjuntos de ítems (sin duplicados)
                    items_filtrados_ids = set(items_con_propiedad_especial).union(items_sin_relacion)

                    # Actualizar la población con los ítems filtrados
                    poblacion = poblacion.filter(idproduct__in=items_filtrados_ids)
                else:
                    # Filtrar por items con el perfil especificado
                    poblacion = poblacion.filter(
                        clases_propiedades__idclase__idclase=1, 
                        clases_propiedades__propiedad=perfil_filtro
                    )

                # Obtener los vínculos relacionados con la clase "Aro" (idclase=2)
                vinculados_aro = tblitemclasevinculo.objects.filter(
                    iditem__in=poblacion, 
                    idclase__idclase=2, 
                    activo=True
                )

                # Obtener aros únicos
                aros = vinculados_aro.values('propiedad').annotate(items_existentes=Count('iditem', distinct=True))
                aros_data = [
                    {"aro": aro["propiedad"] if aro["propiedad"] != "-" else "(-)", 
                    "items_existentes": aro["items_existentes"]}
                    for aro in aros
                ]

                # Agregar los items sin vinculación o con propiedad "-"
                items_sin_aro = poblacion.exclude(
                    clases_propiedades__idclase__idclase=2
                ).count()
                if items_sin_aro > 0:
                    aros_data.append({"aro": "(-)", "items_existentes": items_sin_aro})
                filtro_dinamico["aro"] = aros_data
            # Respuesta final
        return Response({
            "filtro general": filtro_general,
            "filtro dinamico": filtro_dinamico
        })

from rest_framework.pagination import PageNumberPagination

class FixedPageNumberPagination(PageNumberPagination):
    page_size = 10  # Fijamos el tamaño de página a 10
from rest_framework.response import Response

class BusquedaDinamicaViewSet(viewsets.ViewSet):
    permission_classes = [AllowAny]  # Permitir acceso sin autenticación
    serializer_class = TblitemSerializer
    pagination_class = FixedPageNumberPagination  # Clase de paginación personalizada

    def create(self, request):
        # Obtener el JSON del cuerpo de la solicitud
        data = request.data
        params = request.query_params
        query = Q()

        # Filtros opcionales
        cadena_busqueda = data.get("cadena_busqueda")
        id_categoria = data.get("id_categoria")
        id_marca_list = data.get("id_marca", [])
        id_modelo_list = data.get("id_modelo", [])
        clase_categoria = data.get("clase_categoria", [])
        ancho_list = data.get("ancho", [])

        # Construcción del query
        if id_categoria:
            query &= Q(categoria_relacionada__idcategoria__id=id_categoria)
        if cadena_busqueda:
            cadena_search_query = (
                Q(codigosku__icontains=cadena_busqueda) |
                Q(titulo__icontains=cadena_busqueda) |
                Q(descripcion__icontains=cadena_busqueda) |
                Q(altura__icontains=cadena_busqueda) |
                Q(ancho__icontains=cadena_busqueda)
            )
            query &= cadena_search_query
        if ancho_list:
            query &= Q(ancho__in=ancho_list)
        if id_marca_list:
            query &= Q(idmodelo__idmarca_id__in=id_marca_list)
        if id_modelo_list:
            query &= Q(idmodelo__in=id_modelo_list)

        # Filtrar ítems por ancho
        poblacion = Tblitem.objects.filter(activo=True).filter(query).distinct()

# Caso especial: Manejo de clase_categoria
        if clase_categoria:
            for clase in clase_categoria:
                id_clase = clase.get("id_clase")
                propiedad_list = clase.get("propiedad", [])

                if id_clase is None or not isinstance(propiedad_list, list):
                    return Response(
                        {'error': 'Cada objeto en "clase_categoria" debe contener "id_clase" y "propiedad" como lista.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Filtro para las propiedades de esta clase
                clase_subquery = Q()

                # Caso especial: propiedad "(-)"
                if "(-)" in propiedad_list:
                    vinculos_con_propiedad_especial = tblitemclasevinculo.objects.filter(
                        idclase__idclase=id_clase,
                        propiedad="-",
                        iditem__in=poblacion.values_list('idproduct', flat=True)
                    ).values_list('iditem', flat=True)

                    # Ítems sin relación con la clase
                    vinculos_de_clase = tblitemclasevinculo.objects.filter(
                        idclase__idclase=id_clase,
                        iditem__in=poblacion.values_list('idproduct', flat=True)
                    ).values_list('iditem', flat=True)

                    items_sin_relacion = poblacion.exclude(idproduct__in=vinculos_de_clase)

                    # Unir ítems con propiedad "-" y sin relación
                    items_filtrados_ids = set(vinculos_con_propiedad_especial).union(
                        items_sin_relacion.values_list('idproduct', flat=True)
                    )
                    clase_subquery |= Q(idproduct__in=items_filtrados_ids)

                    # Remover "(-)" para evitar duplicar condiciones
                    propiedad_list = [prop for prop in propiedad_list if prop != "(-)"]

                # Filtro para propiedades restantes
                if propiedad_list:
                    clase_subquery |= Q(
                        clases_propiedades__idclase=id_clase,
                        clases_propiedades__propiedad__in=propiedad_list
                    )

                # Aplicar el subquery de la clase al query principal
                poblacion = poblacion.filter(clase_subquery)

        # Filtrar los ítems finales
        items = poblacion.distinct()

        # Ordenamiento
        ordering = params.get("ordering", None)
        if ordering:
            items = items.order_by(*ordering.split(","))


        # Paginación
        paginator = self.pagination_class()
        paginated_items = paginator.paginate_queryset(items, request)

        # Serialización
        serializer = TblitemSerializer(paginated_items, many=True)

        # Respuesta
        return paginator.get_paginated_response(serializer.data)


class BusqussedaDinamicaViewSet(viewsets.ViewSet):
    permission_classes = [AllowAny]  # Permitir acceso sin autenticación

    def create(self, request):
        # Leer el cuerpo de la solicitud
        data = request.data
        ancho_filtro = data.get('ancho', [])
        clase_categoria = data.get('clase_categoria', [])

        # Validar que "ancho" sea una lista no vacía
        if not ancho_filtro:
            return Response(
                {'error': 'El campo "ancho" es obligatorio y debe ser una lista no vacía.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Filtrar los ítems por ancho
        poblacion = Tblitem.objects.filter(activo=True, ancho__in=ancho_filtro)
        print(poblacion.count() , "aqui")

        # Manejo de caso especial
        resultado = []
        for clase in clase_categoria:
            id_clase = clase.get('id_clase')
            propiedad_filtro = clase.get('propiedad', [])

            if id_clase is None or not isinstance(propiedad_filtro, list):
                return Response(
                    {'error': 'Cada objeto en "clase_categoria" debe contener "id_clase" y "propiedad" como lista.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Filtro para caso especial: propiedad = "(-)" o sin relación con la clase
            if "(-)" in propiedad_filtro:
                vinculos_con_propiedad_especial = tblitemclasevinculo.objects.filter(
                    idclase__idclase=id_clase,
                    propiedad="-",
                    iditem__in=poblacion.values_list('idproduct', flat=True)
                ).values_list('iditem', flat=True)

                # 2. Identificar los ítems que no tienen ninguna relación con la clase
                vinculos_de_clase = tblitemclasevinculo.objects.filter(
                    idclase__idclase=id_clase,
                    iditem__in=poblacion.values_list('idproduct', flat=True)
                ).values_list('iditem', flat=True)

                # Ítems de la población que no están relacionados con la clase
                items_sin_relacion = poblacion.exclude(idproduct__in=vinculos_de_clase)

                # 3. Unir los ítems relacionados con "-" y los que no tienen relación
                items_filtrados_ids = set(vinculos_con_propiedad_especial).union(
                    items_sin_relacion.values_list('idproduct', flat=True)
                )
                items_filtrados = poblacion.filter(idproduct__in=items_filtrados_ids)
                print(items_filtrados.count() , "aqui")
                
            else:
                # Filtrar por propiedades específicas si no es caso especial
                clase_vinculos_especiales = tblitemclasevinculo.objects.filter(
                    idclase__idclase=id_clase,
                    propiedad__in=propiedad_filtro,
                    iditem__in=poblacion.values_list('idproduct', flat=True)
                )

            # Obtener los ítems correspondientes al filtro actual
            #items_especiales = poblacion.filter(
            #    idproduct__in=clase_vinculos_especiales.values_list('iditem', flat=True)
            #)

            # Añadir los ítems encontrados al resultado final
            resultado.extend(poblacion.values('idproduct', 'codigosku', 'titulo', 'ancho', 'stock'))

        # Devolver los resultados organizados
        return Response({'items': resultado}, status=status.HTTP_200_OK)
    

from knox.views import LoginView as KnoxLoginView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import login
from .serializers import CustomAuthTokenSerializer
from .models import CustomUser
from .serializers import CustomUserSerializer
from rest_framework.response import Response
from rest_framework import status, permissions
from django.contrib.auth import login
from rest_framework.exceptions import ValidationError
from knox.views import LoginView as KnoxLoginView

class LoginView(KnoxLoginView):
    permission_classes = (permissions.AllowAny,)

    def post(self, request, format=None):
        # Inicializar el serializer con los datos enviados
        serializer = CustomAuthTokenSerializer(data=request.data)

        # Intentar validar el serializer
        if not serializer.is_valid():
            # Construir una respuesta de error uniforme en JSON
            errors = serializer.errors
            return JsonResponse(
                {
                    "success": False,
                    "error": {
                        "code": "invalid_data",
                        "message": "Se encontraron errores en los datos enviados.",
                        "details": errors,  # Esto incluye los errores del serializer
                    }
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Recuperar el usuario autenticado desde el serializer
        user = serializer.validated_data["user"]

        # Iniciar sesión
        login(request, user)

        # Obtener la respuesta estándar de Knox
        response = super(LoginView, self).post(request, format=None)
        expiry = response.data.get("expiry")
        # Serializar la información del usuario
        user_serializer = CustomUserSerializer(user)

        # Responder con un JSON que combine el token y la información del usuario
        return JsonResponse(
            {   
                "success": True,
                "expiry": expiry,
                "token": response.data["token"],
                "user": user_serializer.data,
            },
            status=status.HTTP_200_OK,
        )
#----------------------------------------------------------------------Login imports
from rest_framework import permissions
from rest_framework.authtoken.serializers import AuthTokenSerializer
from knox.views import LoginView as KnoxLoginView
from django.contrib.auth import login
class LoginViewm(KnoxLoginView):
    
    permission_classes = (permissions.AllowAny,)
    @swagger_auto_schema( request_body=AuthTokenSerializer)

    def post(self, request, format=None):
        print(request)
        serializer = AuthTokenSerializer(data=request.data)
        print(serializer)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        login(request, user)
        return super(LoginView, self).post(request, format=None)



#
class UserInfoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        serializer = CustomUserSerializer(user)
        return Response(serializer.data, status=status.HTTP_200_OK)

class UserPedidosView(ModelViewSet):
    """
    ViewSet para manejar pedidos con soporte para búsquedas, filtros y ordenamiento.
    """
    permission_classes = [IsAuthenticated]
    #queryset = Tblpedido.objects.filter(activo=True).order_by('pk')  # Filtra solo pedidos activos
    serializer_class = TblpedidoSerializer

    # Configuración de filtros, búsquedas y ordenamiento
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['idcliente__nombreusuario', 'total', 'estado']
 # Permite ordenar por fecha o precio
    filterset_class = TblpedidoFilter  # Usar un filtro personalizado
    
    

    def get_queryset(self):
        # Filtrar los pedidos por el usuario autenticado
        user = self.request.user
        return Tblpedido.objects.filter(idcliente=user, activo=True).order_by('pk')
    
      
    
class RegisterAPI(generics.GenericAPIView):
    permission_classes = (permissions.AllowAny,)
    
    serializer_class = RegisterSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        return Response({
            "user": serializer.data,
            
            "token": AuthToken.objects.create(user)[1]
        })

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.mail import send_mail
import uuid
from django.contrib.auth.tokens import PasswordResetTokenGenerator

class PasswordResetRequestView(APIView):
    def post(self, request):
        serializer = PasswordResetRequestSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            user = CustomUser.objects.get(email=email)

            # Generar un token único
            token_generator = PasswordResetTokenGenerator()
            token = token_generator.make_token(user)
            # Enviar correo
            scheme = request.scheme  # 'http' o 'https'
            host = request.get_host()  # Dominio del servidor
            reset_link = f"{scheme}://{host}/ecommerce/password-reset?user_id={user.id}&token={token}"
            send_mail(
                "Recuperación de contraseña",
                f"Usa este enlace para restablecer tu contraseña: {reset_link}",
                "no-reply@example.com",
                [email]
            )
            return Response({"message": "Correo de recuperación enviado."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
from django.contrib.auth.tokens import PasswordResetTokenGenerator

class PasswordResetView(APIView):
    def get(self, request):
        # Obtener parámetros de la URL
        user_id = request.query_params.get("user_id")
        token = request.query_params.get("token")

        # Validar los parámetros
        if not user_id or not token:
            return Response({"error": "Faltan parámetros en la URL."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({"error": "Usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        token_generator = PasswordResetTokenGenerator()
        if not token_generator.check_token(user, token):
            return Response({"error": "Token inválido o expirado."}, status=status.HTTP_400_BAD_REQUEST)

  # Incrustar el HTML básico
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Restablecer Contraseña</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 50px;
                    padding: 20px;
                    text-align: center;
                }}
                form {{
                    max-width: 400px;
                    margin: auto;
                }}
                input {{
                    width: 100%;
                    padding: 10px;
                    margin: 10px 0;
                    box-sizing: border-box;
                }}
                button {{
                    padding: 10px 20px;
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    cursor: pointer;
                }}
            </style>
        </head>
        <body>
            <h1>Restablecer Contraseña</h1>
            <form method="POST" action="/ecommerce/login/">
                <input type="hidden" name="user_id" value="{user_id}">
                <input type="hidden" name="token" value="{token}">
                <label for="new_password">Nueva Contraseña:</label>
                <input type="password" id="new_password" name="new_password" required>
                <label for="confirm_password">Confirmar Contraseña:</label>
                <input type="password" id="confirm_password" name="confirm_password" required>
                <button type="submit">Restablecer</button>
            </form>
        </body>
        </html>
        """
        return HttpResponse(html_content, content_type="text/html")
        # Renderizar el formulario de restablecimiento de contraseña
        #return render(request, "password_reset_form.html", {"user_id": user_id, "token": token})
class AdministracionViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    queryset = Administracion.objects.filter(activo=True).order_by('pk')
    serializer_class = AdministracionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['nombreempresa', 'ruc', 'telefono', 'idmoneda__nombre']
    filterset_fields = ['activo', 'nombreempresa', 'ruc', 'telefono', 'igv', 'idmoneda_id', 'idmoneda__nombre', 'fechacreacion', 'fechamodificacion']



class CuponViewSet(ModelViewSet):
    permission_classes = [AllowRetrieveWithoutAuth] 
    queryset = Cupon.objects.filter(activo=True).order_by('pk')
    serializer_class = CuponSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['codigo', 'estado']
    filterset_fields = {
    'activo': ['exact'],
    'idcupon': ['exact'],
    'cantidaddescuento': ['exact'],
    'estado': ['exact'],
    'fechavigencia': ['exact'],
    'fechacreacion': ['exact'],
    'fechamodificacion': ['exact'],
    'codigo': ['exact'],  # Búsqueda literal
}
class MarcaViewSet(ModelViewSet):
    queryset = Marca.objects.filter(activo=True).order_by('pk')
    serializer_class = MarcaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre']
    filterset_fields = ['activo', 'id', 'nombre']


class TblreclamacionViewSet(ModelViewSet):
    permission_classes = [AllowPostWithoutAuth]
    queryset = Tblreclamacion.objects.filter(activo=True).order_by('pk')
    serializer_class = TblreclaisionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['nombre']
    filterset_fields = '__all__'



class MarcaViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    queryset = Marca.objects.filter(activo=True).order_by('pk')
    serializer_class = MarcaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre']
    filterset_fields = ['activo', 'id', 'nombre']

class TblcategoriaViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    queryset = Tblcategoria.objects.filter(activo=True).order_by('pk')
    serializer_class = TblcategoriaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre']
    filterset_fields = ['activo', 'id', 'nombre']

class FleteViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Flete.objects.filter(activo=True).order_by('pk')
    serializer_class = FleteSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['iddepartamento',"idcategoria__nombre"]
    filterset_fields = ['precio','activo', 'id','idcategoria_id', 'idcategoria__nombre',"iddepartamento"]

    @action(detail=False, methods=['put'], url_path='matriz-update', serializer_class=DepartamentoSerializer)
    def matriz_actualizacion(self, request, *args, **kwargs):
        """
        Actualizar los registros de un departamento.
        """
        serializer = self.get_serializer(data=request.data, many=True)
        if serializer.is_valid():
            # Actualizar registros
            for departamento_data in serializer.validated_data:
                iddepartamento = departamento_data['iddepartamento']
                valores = departamento_data['valores']

                # Eliminar registros existentes del departamento
                Flete.objects.filter(iddepartamento=iddepartamento).delete()

                # Crear nuevos registros
                fletes = [
                    Flete(
                        iddepartamento=iddepartamento,
                        idcategoria_id=valor['idcategoria'],
                        precio=valor['val']
                    )
                    for valor in valores
                ]
                Flete.objects.bulk_create(fletes)

            return Response({'message': 'Fletes actualizados exitosamente.'}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'], url_path='matriz-lista')
    def matriz_lista(self, request, *args, **kwargs):
        """
        Listar los departamentos con sus categorías y precios.
        """
        departamentos = Flete.objects.values('iddepartamento').distinct()
        data = []

        for departamento in departamentos:
            iddepartamento = departamento['iddepartamento']
            valores = Flete.objects.filter(iddepartamento=iddepartamento).values(
                categoria=F('idcategoria__id'),
                val=F('precio')
            )
            data.append({
                'iddepartamento': iddepartamento,
                'valores': list(valores)
            })

        return Response(data, status=status.HTTP_200_OK)
    @swagger_auto_schema(
        operation_description="Creación masiva de Fletes",
        request_body=DepartamentoSerializer(many=True),
        responses={201: "Fletes creados exitosamente", 400: "Error en los datos proporcionados"}
    )
    @action(detail=False, methods=['post'], url_path='matriz-creacion', serializer_class=DepartamentoSerializer)
    def matriz_creacion(self, request):
        """
        Acción personalizada para crear múltiples registros de Flete.
        """
        serializer = DepartamentoSerializer(data=request.data, many=True)
        if serializer.is_valid():
            serializer.save()  # Llama al método `create` en el serializer
            return Response({'message': 'Fletes creados exitosamente.'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class TblmodeloViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    queryset = Tblmodelo.objects.filter(activo=True).order_by('pk')
    serializer_class = TblmodeloSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre']
    filterset_fields = ['activo', 'id', 'nombre', 'idmarca_id']




class MonedaViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Moneda.objects.filter(activo=True).order_by('pk')
    serializer_class = MonedaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['nombre']
    filterset_fields = ['activo', 'idmoneda', 'nombre', 'estado', 'fechacreacion', 'fechamodificacion']



class PromocionViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Promocion.objects.filter(activo=True).order_by('pk')
    serializer_class = PromocionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = [ 'estado']
    filterset_fields = ['activo', 'idpromocion', 'estado', 'fechacreacion', 'fechamodificacion']


class TblcarritoViewSet(ModelViewSet):
    queryset = Tblcarrito.objects.filter(activo=True).order_by('pk')
    serializer_class = TblcarritoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['idusuario__nombreusuario', 'preciototal']
    filterset_fields = ['activo', 'idusuario_id', 'preciototal']



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from requests.auth import HTTPBasicAuth
import base64
class AuthCredentialsView(APIView):
    """
    Servicio para generar el encabezado 'Authorization' con la autenticación básica.
    """

    def get(self, request, *args, **kwargs):
        # Obtener las credenciales desde la configuración de Django o desde los parámetros
        username = settings.IZIPAY_USERNAME  # O puedes recibirlo como parámetro en la URL si es necesario
        password = settings.IZIPAY_PASSWORD  # O puedes recibirlo como parámetro también

        # Concatenar el nombre de usuario y la contraseña con un ':' entre ellos
        user_pass_string = f"{username}:{password}"

        # Codificar en base64
        base64_encoded = base64.b64encode(user_pass_string.encode('utf-8')).decode('utf-8')

        # Construir el encabezado de autorización
        auth_header = f"Basic {base64_encoded}"

        # Retornar la respuesta con el encabezado generado
        return Response(
            data={
                "authorization_header": auth_header
            },
            status=status.HTTP_200_OK
        )

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets
from .models import Tblitem
from .serializers import TblitemSerializer
from .filters import TblitemclasepropiedadFilter
from .filters import *
from rest_framework import filters


from django.db import transaction
from django.core.exceptions import ValidationError
from django.db.models import Sum, F, DecimalField
from django.utils.timezone import now 

from openpyxl.styles import Alignment, Font, PatternFill,Border, Side
from openpyxl import Workbook
from django.http import HttpResponse

from django.db.models import OuterRef, Subquery, Max
from django.forms.models import model_to_dict
from openpyxl.utils import get_column_letter
from django.http import StreamingHttpResponse
from openpyxl import Workbook
import csv

from io import BytesIO
from django.db.models import Prefetch
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
class ExcelStreamingResponse:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Plantilla"

    def write_headers(self, headers):
        self.ws.append(headers)

    def write_row(self, row):
        self.ws.append(row)

    def generate(self):
        from io import BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        yield output.read()
        
        
import pandas as pd
from collections import Counter
class TblitemViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Tblitem.objects.filter(activo=True).prefetch_related(
        'clases_propiedades',
        'clases_propiedades__idclase'
        # 'clases_propiedades__propiedad'
    ).all()
    serializer_class = TblitemSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    filterset_class = TblitemFilter
    search_fields = [
        'codigosku',       # Buscamos en el SKU
        'descripcion',     # Buscamos en la descripción
        #'clases_propiedades__idclase__nombre', # Buscamos en el nombre de la clase
        #'clases_propiedades__propiedad',
        'categoria_relacionada__idcategoria__nombre',
        'stock',
        'titulo',
        'precionormal'
        # Buscamos en la propiedad
    ]
  
    #----------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='filtrobusqueda')
    def busqueda_dinamicaitems(self, request):
        """
        Acción personalizada para realizar búsquedas dinámicas con filtros adicionales.
        """
        data = request.query_params
        query = Q()

        # Filtros adicionales opcionales
        cadena_busqueda = data.get("cadena_busqueda")
        id_categoria = data.get("id_categoria")
        id_marca_list = data.getlist("id_marca", [])
        id_modelo_list = data.getlist("id_modelo", [])
        clase_categoria = data.getlist("clase_categoria", [])

        # Aplicar los filtros si están presentes
        if id_categoria:
            query &= Q(categoria_relacionada=id_categoria)

        if cadena_busqueda:
            query &= Q(titulo__icontains=cadena_busqueda)

        if id_marca_list:
            query &= Q(idmodelo__idmarca_id__in=id_marca_list)

        if id_modelo_list:
            query &= Q(idmodelo__in=id_modelo_list)

        if clase_categoria:
            subquery = Q()
            for item in clase_categoria:
                id_clase = item.get("id_clase")
                propiedad_list = item.get("propiedad", [])

                if id_clase and propiedad_list:
                    subquery |= Q(
                        clases_propiedades__idclase=id_clase,
                        clases_propiedades__propiedad__in=propiedad_list
                    )
            query &= subquery

        # Obtener queryset filtrado
        items = Tblitem.objects.filter(activo=True).filter(query).distinct()
        
         # Aplicar ordenamiento manual
        ordering = data.get("ordering")
        if ordering:
            items = items.order_by(*ordering.split(","))
            
        # Serializar los resultados (la paginación y el ordenamiento se aplicarán automáticamente)
        paginator = self.pagination_class()
        paginated_items = paginator.paginate_queryset(items, request, view=self)

        # Serializar los resultados
        serializer = self.get_serializer(paginated_items, many=True)

        # Responder con los datos paginados
        return paginator.get_paginated_response(serializer.data)
    #----------------------------------------------------------------

    @swagger_auto_schema(
        operation_description="Obtiene el detalle completo del ítem junto con el número de pedidos y los ingresos totales.",
        responses={
            200: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "item_data": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            #**{
                            #    field: serializer_field.schema for field, serializer_field in TblitemSerializer().fields.items()
                            #},
                            "numero_pedidos": openapi.Schema(
                                type=openapi.TYPE_INTEGER,
                                description="Número total de pedidos en los que aparece este producto."
                            ),
                            "ingresos_totales": openapi.Schema(
                                type=openapi.TYPE_NUMBER,
                                format=openapi.FORMAT_DECIMAL,
                                description="Ingresos totales generados por este producto (suma del precio unitario multiplicado por la cantidad)."
                            ),
                        },
                    ),
                },
            ),
            404: openapi.Response('Error: El ítem no existe.'),
        },
    )
    @action(detail=True, methods=['get'], url_path='detalles-ventas')
    def detalles_ventas(self, request, pk=None):
        """
        Obtiene el detalle completo del ítem junto con el número de pedidos y los ingresos totales.
        """
        try:
            item = Tblitem.objects.filter(activo=True).get(pk=pk)
        except Tblitem.DoesNotExist:
            return Response({'error': 'El ítem no existe.'}, status=404)

        ## quierosparar los pediodo que encuentren ese item por moneda que eso esta en su idpedido
        ## y que se sumen los montos de preciototal  pero por moneda uq equeden separado tambien
        ## tener el administracion id= 1 y sacar la moneda
        ## tener el valor de ultimo tipocambio activo  de la moneda de administracion
        ##  de las sumas de preciototal separados por moneda  que no coincidan con la moneda de adminisracion
          ## que q esas sumas que no coincidad con el idmoneda entonces sean convertidos
          ## convertidos por el valor deltipode cambio de la moneda 
        # Filtrar los pedidos que incluyen este ítem
       
        # O si prefieres hacer una verificación más explícita, puedes realizar una subconsulta
        # que solo tome los detalles de pedidos cuyo idpedido tiene un idtransaccion no nulo
        detalles_pedidos_validos = Tbldetallepedido.objects.filter(
            activo=True,
            idproduct=item,
            idpedido__idtransaccion__isnull=False,
            idpedido__activo=True,
            idpedido__estado__in=[2, 4]
        )
        detalles_pedidos = detalles_pedidos_validos
        
        numero_pedidos = detalles_pedidos.count()
                # Obtener la moneda de la administración
        try:
            admin = Administracion.objects.filter(activo=True).get(pk=1)
            moneda_administracion = admin.idmoneda
        except Administracion.DoesNotExist:
            return Response({'error': 'No se encontró la configuración de administración.'}, status=500)


        
        
        

        #from django.db.models import Sum, F, OuterRef, Subquery

        # Obtener los ingresos por moneda (suma de ganancias)
        ingresos_por_moneda = detalles_pedidos.values('idpedido__idmoneda').annotate(
            total_ganancias=Sum(F('preciototal'))
        )
        print(ingresos_por_moneda)

        # Obtener los tipos de cambio actuales para las monedas involucradas
        monedas_ids = [ingreso['idpedido__idmoneda'] for ingreso in ingresos_por_moneda]

        # Subconsulta para obtener la última fecha registrada de cada moneda
        ultima_fecha_subquery = Tipocambio.objects.filter(
            idmoneda=OuterRef('idmoneda'),
            idmoneda__in=monedas_ids
        ).order_by('-fecha').values('fecha')[:1]

        # Filtrar los tipos de cambio usando la subconsulta
        tipos_cambio = Tipocambio.objects.filter(
            idmoneda__in=monedas_ids,
            fecha=Subquery(ultima_fecha_subquery)
        )

        # Crear un diccionario de tipos de cambio (moneda -> tipo_cambio)
        tipos_cambio_dict = {tipo.idmoneda.idmoneda: tipo.tipocambio for tipo in tipos_cambio}
    
        # Inicializar el total de ingresos
        total_ingresos_convertidos = 0
        total_ingresos_sin_conversion = 0  # Ingresos en moneda de administración sin conversión

        # Primer paso: Sumar los ingresos en la moneda original y convertir si es necesario
        for ingreso in ingresos_por_moneda:
            idmoneda = ingreso['idpedido__idmoneda']
            total_ganancias = ingreso['total_ganancias']

            # Si la moneda del pedido es la misma que la moneda de administración, sumamos directamente
            if idmoneda == moneda_administracion.idmoneda:
                total_ingresos_sin_conversion += total_ganancias
                #total_ingresos_convertidos += Decimal(total_ganancias) / tipo_cambio
            else:
                # Si la moneda es diferente, convertimos las ganancias a la moneda de administración
                if idmoneda in tipos_cambio_dict:
                    tipo_cambio = tipos_cambio_dict[idmoneda]
                    #total_ingresos_convertidos += total_ganancias / tipo_cambio
                    total_ingresos_convertidos += Decimal(total_ganancias) * tipo_cambio
            
        # Obtener el tipo de cambio de la moneda de administración
        tipo_cambio_administracion = tipos_cambio_dict.get(moneda_administracion.idmoneda, 1)

        # Convertir el total acumulado a la moneda de administración
        total_final = total_ingresos_convertidos * tipo_cambio_administracion + Decimal(total_ingresos_sin_conversion)  # Convertir a Decimal


        
        
        
        
        serializer = TblitemSerializer(item)

        # Agregar información adicional al diccionario serializado
        item_data = serializer.data
        item_data['numero_pedidos']= numero_pedidos
        item_data['total_ingresos'] = round(total_final, 2)  # Suponiendo que total_final es calculado antes
        item_data['moneda_administracion'] = moneda_administracion.nombre  # Asumiendo que tienes esta variable

        # Retornar la respuesta con el detalle completo y la información adicional
        return Response({
            'item': item_data
        })
    
    #----------------------------------------------------------------
  
    @action(detail=True, methods=['patch'], url_path='update', serializer_class=TblitemTestSerializer)
    @transaction.atomic
    def itempartial_update(self, request, pk=None):
        """
        Actualiza parcialmente un item identificado por su ID en el URL.
        """
        
        try:
            # Obtener el item por medio del ID en el URL
            item = Tblitem.objects.get(pk=pk)
            
        except Tblitem.DoesNotExist:
            return Response({"error": "El item no existe."}, status=status.HTTP_404_NOT_FOUND)
            

        # Validar los datos enviados (parciales)
        serializer = self.get_serializer(item, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        

        try:
            with transaction.atomic():
                # Actualizar campos proporcionados
                item_data = validated_data.get('item', {})
                for key, value in item_data.items():
                    setattr(item, key, value)
                item.save()

                # Actualizar relaciones selectivamente
                if 'vinculos' in validated_data:
                    self.patch_item_vinculos(item, json.loads(validated_data.get('vinculos', '[]')))
                if 'categorias' in validated_data:
                    cate= json.loads(validated_data.get('categorias', '[]'))
                    self.patch_item_categorias(item, cate)
                if 'cupones' in validated_data:
                    self.patch_item_cupones(item, json.loads(validated_data.get('cupones', '[]')))
                if 'itemsrelacionados' in validated_data:

                    self.patch_item_itemsrelacionados(item, json.loads(validated_data.get('itemsrelacionados', '[]')))

                # Actualizar imágenes principales y adicionales
                if 'imagenprincipal' in validated_data:
                    item.imagenprincipal = validated_data['imagenprincipal']
                    item.save()
                
                if 'idmodelo' in validated_data:
                    try:
                        modelo_instance = Tblmodelo.objects.get(id=validated_data['idmodelo'])
                        item.idmodelo = modelo_instance # Asignar la instancia de Tblmodelo
                    except Tblmodelo.DoesNotExist:
                        raise ValueError(f"El modelo con ID {item_data['idmodelo']} no existe.") # Asignar la imagen principal


                if 'imagenes_eliminartodas' in validated_data and validated_data['imagenes_eliminartodas']:
    # Caso 1: Si se especifica `imagenes_eliminartodas=True`, eliminar todas las imágenes
                    Tblimagenitem.objects.filter(idproduct=item).delete()

                elif 'imagenes' in validated_data:
                    imagenes_data = validated_data.get('imagenes', [])

                    if imagenes_data:
                        # Caso 2: Si `imagenes` contiene imágenes, eliminar las existentes y añadir las nuevas
                        Tblimagenitem.objects.filter(idproduct=item).delete()
                        for imagen in imagenes_data:
                            if imagen:  # Verifica que la imagen no esté vacía
                                Tblimagenitem.objects.create(
                                    idproduct=item,
                                    imagen=imagen,
                                    estado=1  # Ajusta según la lógica de tu modelo
                                )
                    else:
                        # Caso 3: Si `imagenes` está presente pero vacío, no hacer nada
                        pass

                # Serializar y responder
                item_serializer = TblitemSerializer(item)
                return Response({
                    "message": "Item actualizado parcialmente con éxito.",
                    "item": item_serializer.data
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"Error inesperado: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


    def patch_item_vinculos(self, item, vinculos_data):
        """
        Actualiza los vínculos asociados a un ítem de forma parcial.
        Elimina los vínculos existentes que no están en los datos enviados.
        """
        # Obtener las relaciones actuales desde la base de datos
        relaciones_actuales = tblitemclasevinculo.objects.filter(iditem=item)
        relaciones_por_id = {rel.id: rel for rel in relaciones_actuales}
        relaciones_por_clase = {rel.idclase_id: rel for rel in relaciones_actuales}

        # Rastrear los vínculos que deben conservarse
        vinculos_a_conservar = set()

        for vinculo in vinculos_data:
            if 'id' in vinculo:  # Relación existente enviada por ID
                if vinculo['id'] in relaciones_por_id:
                    relacion = relaciones_por_id[vinculo['id']]
                    # Actualizar solo si los datos han cambiado
                    if relacion.idclase_id != vinculo.get('idclase') or relacion.propiedad != vinculo.get('propiedad'):
                        relacion.idclase_id = vinculo.get('idclase', relacion.idclase_id)
                        relacion.propiedad = vinculo.get('propiedad', relacion.propiedad)
                        relacion.save()
                    vinculos_a_conservar.add(relacion.id)

            elif 'idclase' in vinculo:  # Relación nueva o existente enviada por clase
                if vinculo['idclase'] in relaciones_por_clase:
                    relacion = relaciones_por_clase[vinculo['idclase']]
                    # Actualizar la propiedad si es necesario
                    if relacion.propiedad != vinculo.get('propiedad', relacion.propiedad):
                        relacion.propiedad = vinculo.get('propiedad', relacion.propiedad)
                        relacion.save()
                    vinculos_a_conservar.add(relacion.id)
                else:  # Crear una nueva relación
                    nueva_relacion = tblitemclasevinculo.objects.create(
                        iditem=item,
                        idclase_id=vinculo['idclase'],
                        propiedad=vinculo.get('propiedad', "")
                    )
                    vinculos_a_conservar.add(nueva_relacion.id)

        # Eliminar vínculos no enviados
        relaciones_actuales.exclude(id__in=vinculos_a_conservar).delete()


    def patch_item_categorias(self, item, categorias_data):
        """
        Actualiza las categorías asociadas a un ítem de forma parcial.
        Elimina las relaciones existentes que no están en los datos enviados.
        """
        # Obtener las relaciones actuales desde la base de datos
        relaciones_actuales = tblitemcategoria.objects.filter(iditem=item)
        relaciones_por_id = {rel.id: rel for rel in relaciones_actuales}
        relaciones_por_categoria = {rel.idcategoria_id: rel for rel in relaciones_actuales}

        # Rastrear las relaciones que deben conservarse
        categorias_a_conservar = set()

        for categoria in categorias_data:
            if 'id' in categoria:  # Relación existente enviada por ID
                if categoria['id'] in relaciones_por_id:
                    relacion = relaciones_por_id[categoria['id']]
                    # Actualizar solo si los datos han cambiado
                    if relacion.idcategoria_id != categoria.get('idcategoria'):
                        relacion.idcategoria_id = categoria.get('idcategoria')
                        relacion.save()
                    categorias_a_conservar.add(relacion.id)

            elif 'idcategoria' in categoria:  # Relación nueva o existente enviada por categoría
                if categoria['idcategoria'] in relaciones_por_categoria:
                    relacion = relaciones_por_categoria[categoria['idcategoria']]
                    categorias_a_conservar.add(relacion.id)
                else:
                    nueva_relacion = tblitemcategoria.objects.create(
                        iditem=item,
                        idcategoria_id=categoria['idcategoria']
                    )
                    categorias_a_conservar.add(nueva_relacion.id)
        
        # Eliminar relaciones no enviadas
        relaciones_actuales.exclude(id__in=categorias_a_conservar).delete()


    def patch_item_cupones(self, item, cupones_data):
        """
        Actualiza los cupones asociados a un ítem de forma parcial.
        Elimina las relaciones existentes que no están en los datos enviados.
        """
        # Obtener las relaciones actuales desde la base de datos
        relaciones_actuales = tblitemcupon.objects.filter(iditem=item)
        relaciones_por_id = {rel.id: rel for rel in relaciones_actuales}
        relaciones_por_cupon = {rel.idcupon_id: rel for rel in relaciones_actuales}

        # Rastrear las relaciones que deben conservarse
        cupones_a_conservar = set()

        for cupon in cupones_data:
            if 'id' in cupon:  # Relación existente enviada por ID
                if cupon['id'] in relaciones_por_id:
                    relacion = relaciones_por_id[cupon['id']]
                    # Actualizar solo si los datos han cambiado
                    if relacion.idcupon_id != cupon.get('idcupon'):
                        relacion.idcupon_id = cupon.get('idcupon')
                        relacion.save()
                    cupones_a_conservar.add(relacion.id)

            elif 'idcupon' in cupon:  # Relación nueva o existente enviada por cupón
                if cupon['idcupon'] in relaciones_por_cupon:
                    relacion = relaciones_por_cupon[cupon['idcupon']]
                    cupones_a_conservar.add(relacion.id)
                else:
                    nueva_relacion = tblitemcupon.objects.create(
                        iditem=item,
                        idcupon_id=cupon['idcupon']
                    )
                    cupones_a_conservar.add(nueva_relacion.id)

        # Eliminar relaciones no enviadas
        relaciones_actuales.exclude(id__in=cupones_a_conservar).delete()


    def patch_item_itemsrelacionados(self, item, itemsrelacionados_data):
        """
        Actualiza los ítems relacionados asociados a un ítem de forma parcial.
        Elimina las relaciones existentes que no están en los datos enviados.
        """
        # Obtener las relaciones actuales desde la base de datos
        relaciones_actuales = Tblitemrelacionado.objects.filter(item=item)
        relaciones_por_id = {rel.id: rel for rel in relaciones_actuales}
        relaciones_por_item_relacionado = {rel.item_relacionado_id: rel for rel in relaciones_actuales}

        # Rastrear las relaciones que deben conservarse
        itemsrelacionados_a_conservar = set()

        for itemsrelacionados in itemsrelacionados_data:
            if 'id' in itemsrelacionados:  # Relación existente enviada por ID
                if itemsrelacionados['id'] in relaciones_por_id:
                    relacion = relaciones_por_id[itemsrelacionados['id']]
                    # Actualizar solo si los datos han cambiado
                    if relacion.item_relacionado_id != itemsrelacionados.get('item_relacionado'):
                        relacion.item_relacionado_id = itemsrelacionados.get('item_relacionado')
                        relacion.save()
                    itemsrelacionados_a_conservar.add(relacion.id)

            elif 'item_relacionado' in itemsrelacionados:  # Relación nueva o existente enviada por ítem relacionado
                if itemsrelacionados['item_relacionado'] in relaciones_por_item_relacionado:
                    relacion = relaciones_por_item_relacionado[itemsrelacionados['item_relacionado']]
                    itemsrelacionados_a_conservar.add(relacion.id)
                else:
                    nuevo_relacionado = Tblitemrelacionado.objects.create(
                        item=item,
                        item_relacionado_id=itemsrelacionados['item_relacionado']
                    )
                    itemsrelacionados_a_conservar.add(nuevo_relacionado.id)

        # Eliminar relaciones no enviadas
        relaciones_actuales.exclude(id__in=itemsrelacionados_a_conservar).delete()

  
    #---------
    @action(detail=False, methods=['post'], url_path='upload-multiple', serializer_class=TblitemTestSerializer)
    @transaction.atomic
    def upload_multiple(self, request):
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data

        # Desempaqueta los datos de entrada
        item_data = validated_data.get('item', {})
        vinculos_data = json.loads(validated_data.get('vinculos', '[]'))
        categorias_data = json.loads(validated_data.get('categorias', '[]'))
        cupones_data = json.loads(validated_data.get('cupones', '[]'))
        itemsrelacionados_data = json.loads(validated_data.get('itemsrelacionados', '[]'))
        imagenes_data = validated_data.get('imagenes', [])
        imagen_principal = validated_data.get('imagenprincipal', None)  # Asignado el campo 'imagenprincipal' para la imagen principal
        idmodelo = validated_data.get('idmodelo', None)  # Asignado el campo 'imagenprincipal' para la imagen principal

        try:
            with transaction.atomic():
                # Crear el item y manejar relaciones
                item_data_dict = item_data.copy()  # Hacer una copia para agregar la imagen principal

                # Si hay una imagen principal, agregarla a los datos del item
                if imagen_principal:
                    item_data_dict['imagenprincipal'] = imagen_principal  # Asignar la imagen principal
                if idmodelo:
                    try:
                        modelo_instance = Tblmodelo.objects.get(id=idmodelo)
                        item_data_dict['idmodelo'] = modelo_instance  # Asignar la instancia de Tblmodelo
                    except Tblmodelo.DoesNotExist:
                        raise ValueError(f"El modelo con ID {item_data_dict['idmodelo']} no existe.") # Asignar la imagen principal

                # Crear el item con los datos (incluyendo imagen principal)
                item = self.create_item_with_vinculos(item_data_dict, vinculos_data)
                self.create_item_categorias(item, categorias_data)
                self.create_item_cupones(item, cupones_data)
                self.create_item_itemsrelacionados(item, itemsrelacionados_data)

                # Procesar imágenes adicionales
                for imagen in imagenes_data:
                    Tblimagenitem.objects.create(
                        idproduct=item,
                        imagen=imagen,
                        estado=1  # Ajusta según tu lógica
                    )

                # Serializa el item creado (usando el serializer correcto para la respuesta)
                item_serializer = TblitemSerializer(item)

                return Response({
                    "message": "Item creado con éxito.",
                    "item": item_serializer.data,  # Serializa el item recién creado
                }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": f"Error inesperado: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    
    def create_item_with_vinculos(self, item_data, vinculos_data):
        """
        Crea un item y sus vínculos asociados en una operación atómica.
        """
        # Crear el item
        item = Tblitem.objects.create(**item_data)

        # Crear los vínculos asociados
        vinculos = [
            tblitemclasevinculo(
                iditem=item,
                idclase_id=vinculo.get("idclase"),
                propiedad=vinculo.get("propiedad", "")
            )
            for vinculo in vinculos_data
        ]
        tblitemclasevinculo.objects.bulk_create(vinculos)

        return item
    def create_item_categorias(self, item, categorias_data):
        """
        Crea las categorías asociadas a un item.
        """
        # Crear las categorías asociadas
        categorias = [
            tblitemcategoria(
                iditem=item,
                idcategoria_id=categoria.get("idcategoria")
            )
            for categoria in categorias_data
        ]
        tblitemcategoria.objects.bulk_create(categorias)


        
    def create_item_cupones(self, item, cupones_data):
        """
        Crea los cupones asociados a un item.
        """
        # Crear los cupones asociados
        cupones = [
            tblitemcupon(
                iditem=item,
                idcupon_id=cupon.get("idcupon")
            )
            for cupon in cupones_data
        ]
        tblitemcupon.objects.bulk_create(cupones)
        
    def create_item_itemsrelacionados(self, item, itemsrelacionados_data):
        """
        Crea los items relacionados a un item.
        """
        # Crear los items relacionados asociados
        items_relacionados = [
            Tblitemrelacionado(
                item=item,
                item_relacionado_id=itemsrelacionados.get("item_relacionado")  # Asegúrate de usar _id si es un ForeignKey
            )
            for itemsrelacionados in itemsrelacionados_data  # Itera correctamente sobre los datos
        ]
        
        # Crear en masa los registros de relaciones
        Tblitemrelacionado.objects.bulk_create(items_relacionados)


    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        try:
            file = request.FILES.get('file')
            if not file:
                return Response({"error": "No se ha proporcionado ningún archivo."},
                                status=status.HTTP_400_BAD_REQUEST)

            df = pd.read_excel(file, header=1, dtype=str, engine="openpyxl")

            required_columns = ["CODIGO(SKU)", "NOMBRE DEL PRODUCTO", "STOCK", "PRECIO", "MARCA", "MODELO", "ESTADO"]
            for column in required_columns:
                if column not in df.columns:
                    return Response({"error": f"Falta la columna requerida: {column}"}, status=status.HTTP_400_BAD_REQUEST)

            # 🔹 Tipos de datos esperados
            expected_types = {
                "CODIGO(SKU)": str,
                "NOMBRE DEL PRODUCTO": str,
                "STOCK": float,  # Convertiremos a número
                "PRECIO": float,  # Convertiremos a número
                "MARCA": int,  # IDs deben ser enteros
                "MODELO": int,  # IDs deben ser enteros
                "ESTADO": int,  # Debe ser un número
                "ANCHO": float,  # Opcional, pero si está debe ser número
                "CATEGORIA": int,  # Opcional, pero si está debe ser número
            }

            # 🔹 Detectar SKUs duplicados en el archivo
            skus_archivo = [str(row["CODIGO(SKU)"]).strip() for _, row in df.iterrows()]
            skus_repetidos = [sku for sku, count in Counter(skus_archivo).items() if count > 1]

            if skus_repetidos:
                return Response({
                    "message": "Carga masiva cancelada. Se encontraron SKUs duplicados en el archivo.",
                    "skus_duplicados": skus_repetidos,
                }, status=status.HTTP_400_BAD_REQUEST)

            errors = []
            validated_data = []

            # 🔹 Validación de cada fila
            for index, row in df.iterrows():
                try:
                    # Extraer valores con limpieza de espacios
                    sku = str(row["CODIGO(SKU)"]).strip()
                    titulo = str(row["NOMBRE DEL PRODUCTO"]).strip()
                    stock = row["STOCK"]
                    precio_normal = row["PRECIO"]
                    marca_id = row["MARCA"]
                    modelo_id = row["MODELO"]
                    estado = row["ESTADO"]
                    ancho = row.get("ANCHO", None)
                    categoria_id = row.get("CATEGORIA", None)

                    # 🚨 Validar Tipos de Datos
                    for col, expected_type in expected_types.items():
                        value = row.get(col)
                        if pd.notna(value):  # Si la celda tiene valor
                            try:
                                if expected_type == int:
                                    row[col] = int(value)
                                elif expected_type == float:
                                    row[col] = float(value)
                                elif expected_type == str:
                                    row[col] = str(value).strip()
                            except ValueError:
                                errors.append({
                                    "fila": index + 2,
                                    "sku": sku,
                                    "columna": col,
                                    "valor": value,
                                    "error": f"El valor '{value}' debe ser {expected_type.__name__}.",
                                })

                    # 🚨 Verificar si el SKU ya existe en la BD
                    if sku in Tblitem.objects.values_list("codigosku", flat=True):
                        errors.append({
                            "fila": index + 2,
                            "sku": sku,
                            "error": "El SKU ya existe en la base de datos.",
                        })
                        continue  # No procesar este SKU

                    # 🚨 Validar existencia en la BD
                    if not Marca.objects.filter(id=marca_id).exists():
                        errors.append({"fila": index + 2, "error": f"La marca con ID {marca_id} no existe."})
                    if not Tblmodelo.objects.filter(id=modelo_id, idmarca=marca_id).exists():
                        errors.append({"fila": index + 2, "error": f"El modelo con ID {modelo_id} no existe o no pertenece a la marca {marca_id}."})
                    if categoria_id and not Tblcategoria.objects.filter(id=categoria_id).exists():
                        errors.append({"fila": index + 2, "error": f"La categoría con ID {categoria_id} no existe."})

                    validated_data.append({
                        "sku": sku,
                        "titulo": titulo,
                        "stock": stock,
                        "precio_normal": precio_normal,
                        "marca_id": marca_id,
                        "modelo_id": modelo_id,
                        "estado": estado,
                        "ancho": None if pd.isna(ancho) else ancho,
                        "categoria_id": categoria_id,
                        "fecha_publicacion": row.get("FECHA PUBLICACION") or now(),
                    })

                except Exception as e:
                    errors.append({
                        "fila": index + 2,
                        "sku": row.get("CODIGO(SKU)", "Desconocido"),
                        "error": str(e),
                    })

            # 🚨 Si hay errores, detener la carga sin tocar la BD
            if errors:
                return Response({
                    "message": "Carga masiva cancelada. Se encontraron errores de validación.",
                    "errors": errors,
                }, status=status.HTTP_400_BAD_REQUEST)

            created_items = []

            # 🔄 **Guardar los datos en la BD dentro de una transacción segura**
            with transaction.atomic():
                for data in validated_data:
                    try:
                        item = Tblitem.objects.create(
                            codigosku=data["sku"],
                            titulo=data["titulo"],
                            precionormal=data["precio_normal"],
                            ancho=data["ancho"],
                            fechapublicacion=data["fecha_publicacion"],
                            destacado=True,
                            nuevoproducto=False,
                            estado=data["estado"],
                            idmodelo=Tblmodelo.objects.get(id=data["modelo_id"]),
                            stock=data["stock"],
                            activo=True,
                        )

                        if data["categoria_id"]:
                            tblitemcategoria.objects.create(
                                iditem=item, idcategoria=Tblcategoria.objects.get(id=data["categoria_id"])
                            )

                        created_items.append(item)

                    except Exception as e:
                        errors.append({
                            "sku": data["sku"],
                            "error": f"Error al insertar en la BD: {str(e)}"
                        })
                        raise  # Revertir la transacción si ocurre un error

            if errors:
                return Response({
                    "message": "Carga masiva completada con errores.",
                    "errors": errors,
                    "items_creados": TblitemSerializer(created_items, many=True).data,
                }, status=status.HTTP_207_MULTI_STATUS)

            return Response({
                "message": "Carga masiva completada con éxito.",
                "items_creados": TblitemSerializer(created_items, many=True).data,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"Error inesperado: {str(e)}"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


    @action(detail=False, methods=['get'], url_path='descargar-plantilla-edicion')
    def descargar_plantilla_edicion(self, request):
        """
        Genera y devuelve una plantilla en Excel con colores, bordes y explicaciones claras.
        """
        # Definir las filas explicativas
        explicaciones = [
            "📌 Identificador único del producto (OBLIGATORIO)",
            "🔢 Cantidad disponible (Si se deja vacío, NO se actualiza)",
            "💰 Precio normal del producto (Si se deja vacío, NO se actualiza)",
            "💲 Precio con descuento (Si se deja vacío, NO se actualiza)"
        ]

        nombres_columnas = ["CODIGO(SKU)", "STOCK", "PRECIO", "PRECIO REBAJADO"]

        valores_permitidos = [
            "Ejemplo: SKU12345",
            "Ejemplo: 50 (Solo números, sin decimales)",
            "Ejemplo: 199.99 (Usar punto para decimales)",
            "Ejemplo: 149.99 (Menor que PRECIO)"
        ]

        # Crear DataFrame con las filas explicativas
        df = pd.DataFrame([explicaciones, nombres_columnas, valores_permitidos])

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_edicion.xlsx'

        # Guardar el DataFrame en el archivo Excel con formato
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name="Plantilla")

            # Obtener el workbook y la hoja activa
            workbook = writer.book
            worksheet = writer.sheets["Plantilla"]

            # Definir formatos de celda
            format_explicaciones = workbook.add_format({'bold': True, 'bg_color': '#FFD966', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
            format_encabezados = workbook.add_format({'bold': True, 'bg_color': '#FFEB9C', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            format_valores = workbook.add_format({'italic': True, 'bg_color': '#D9EAD3', 'border': 1, 'align': 'center', 'valign': 'vcenter'})

            # Aplicar formatos a las filas
            worksheet.set_row(0, 30, format_explicaciones)  # Primera fila: explicaciones
            worksheet.set_row(1, 20, format_encabezados)    # Segunda fila: nombres de columnas
            worksheet.set_row(2, 20, format_valores)        # Tercera fila: valores permitidos

            # Ajustar ancho de columnas
            worksheet.set_column("A:D", 30)

        return response
    
    
    
    @action(detail=False, methods=['post'], url_path='bulk-edit')
    @transaction.atomic
    def bulk_update(self, request):
        """
        Edición masiva de productos desde un archivo Excel.
        - Actualiza STOCK, PRECIO y PRECIO REBAJADO solo si tienen valores.
        - Activa los SKUs presentes en el archivo.
        - Desactiva los SKUs no incluidos en el archivo.
        """
        try:
            file = request.FILES.get('file')
            if not file:
                return Response({"error": "No se ha proporcionado ningún archivo."},
                                status=status.HTTP_400_BAD_REQUEST)

            # Leer archivo
            df = pd.read_excel(file, header=1)

            # Validar columnas requeridas
            required_columns = ["CODIGO(SKU)", "STOCK", "PRECIO", "PRECIO REBAJADO"]
            for column in required_columns:
                if column not in df.columns:
                    return Response(
                        {"error": f"Falta la columna requerida: {column}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            # Obtener SKUs existentes
            skus_existentes = Tblitem.objects.values_list("codigosku", flat=True)
            skus_en_archivo = set()
            errors = []
            updated_items = []

            for index, row in df.iterrows():
                sku = row["CODIGO(SKU)"]
                stock = row.get("STOCK", None)
                precio = row.get("PRECIO", None)
                precio_rebajado = row.get("PRECIO REBAJADO", None)

                if not sku:
                    errors.append({"fila": index + 2, "error": "El SKU es obligatorio."})
                    continue

                try:
                    item = Tblitem.objects.get(codigosku=sku)
                except Tblitem.DoesNotExist:
                    errors.append({"fila": index + 2, "sku": sku, "error": "El SKU no existe. Debe ser creado previamente."})
                    continue

                # Solo actualizar si los valores no son NaN
                update_data = {"estado": 1}  # Asegurar que el SKU se mantenga activo
                if pd.notna(stock) and isinstance(stock, (int, float)) and stock >= 0:
                    update_data["stock"] = stock
                if pd.notna(precio) and isinstance(precio, (int, float)) and precio >= 0:
                    update_data["precionormal"] = precio
                if pd.notna(precio_rebajado) and isinstance(precio_rebajado, (int, float)) and precio_rebajado >= 0:
                    update_data["preciorebajado"] = precio_rebajado

                # Aplicar cambios si hay algo que actualizar
                for key, value in update_data.items():
                    setattr(item, key, value)
                item.save()
                updated_items.append(item)

                skus_en_archivo.add(sku)

            # Desactivar SKUs que no están en el archivo
            skus_a_desactivar = set(skus_existentes) - skus_en_archivo
            cantidad_desactivados = Tblitem.objects.filter(codigosku__in=skus_a_desactivar).update(estado=0)

            # Respuesta final
            if errors:
                return Response({
                    "message": "Edición masiva completada con errores.",
                    "errors": errors,
                    "items_actualizados": TblitemSerializer(updated_items, many=True).data,
                    "cantidad_desactivados": cantidad_desactivados,
                }, status=status.HTTP_207_MULTI_STATUS)

            return Response({
                "message": "Edición masiva completada con éxito.",
                "items_actualizados": TblitemSerializer(updated_items, many=True).data,
                "cantidad_desactivados": cantidad_desactivados,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"Error inesperado: {str(e)}"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """
        Descarga una plantilla Excel mejorada para la carga masiva de productos.
        """
        # Crear el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"

        # Definir las columnas con sus descripciones y ejemplos
        columns = [
            ("CODIGO(SKU)", "SKU123456", "Obligatorio. Identificador único del producto."),
            ("ESTADO", "1", "Obligatorio. Numero que representa el estadoactual del producto."),
            ("NOMBRE DEL PRODUCTO", "Llanta deportiva 17", "Obligatorio. Nombre del producto."),
            ("STOCK", "100", "Obligatorio. OBS : Caso especial: en caso de dejar vacio no hara modificacion, Cantidad disponible."),
            ("PRECIO", "59.99", "Obligatorio. Precio del producto (en formato numérico)."),
            ("MARCA", "1", "Obligatorio. ID de la marca (debe existir en el sistema)."),
            ("MODELO", "30", "Obligatorio. ID del modelo asociado a la marca."),
            ("CATEGORIA", "9", "Obligatorio. ID de la categoría del producto."),
            ("ANCHO", "225", "Obligatorio. Ancho del producto (en cm)."),
            ("PLIEGUES", "4", "Opcional. Número de pliegues."),
            ("IC/IV", "94V", "Opcional. Índice de carga y velocidad."),
            ("APLICACIÓN", "Deportivo", "Opcional. Uso o aplicación del producto."),
            ("SERVICIO", "Turismo", "Opcional. Tipo de servicio al que pertenece."),
            ("ARO", "17", "Opcional. Tamaño del aro."),
            ("ARO PERMITIDO", "16-18", "Opcional. Rango permitido del aro."),
            ("PERFIL", "45", "Opcional. Perfil del producto."),
            ("PRESENTACION", "Caja", "Opcional. Forma de presentación (e.g., caja, unidad)."),
            ("RANGO VELOCIDAD", "240 km/h", "Opcional. Rango de velocidad permitido."),
            ("RUNFLAT", "Sí", "Opcional. Tecnología Runflat (Sí o No)."),
            ("INDICE DE CARGA", "94", "Opcional. Índice de carga."),
        ]

        # Configurar estilos
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        mandatory_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        optional_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_font = Font(bold=True)
        description_font = Font(color="808080", italic=True)
        border_style = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Escribir las filas de encabezados, ejemplos y descripciones
        headers = [column[0] for column in columns]
        examples = [column[1] for column in columns]
        descriptions = [column[2] for column in columns]
        
        ws.append(descriptions)  # Primera fila: Descripciones
        ws.append(headers)  # Segunda fila: Encabezados
        ws.append(examples)  # Tercera fila: Ejemplos

        # Aplicar estilos a las filas
        for col_idx, (header, description) in enumerate(zip(headers, descriptions), start=1):
            # Estilo para descripciones
            description_cell = ws.cell(row=1, column=col_idx)
            description_cell.font = description_font
            description_cell.alignment = Alignment(horizontal="left", vertical="center")
            description_cell.border = border_style

            # Estilo para encabezados
            header_cell = ws.cell(row=2, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = border_style

            # Aplicar color a las columnas obligatorias u opcionales
            is_mandatory = "Obligatorio" in description
            header_cell.fill = mandatory_fill if is_mandatory else optional_fill

        # Ajustar el ancho de las columnas automáticamente
        for col in ws.columns:
            max_length = max(len(str(cell.value) or "") for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Preparar la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_carga_masiva_horizontal.xlsx"'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'], url_path='export-products')
    def export_products(self, request):
        """
        Exportar los productos con sus vínculos a las clases y propiedades en un archivo Excel.
        """
        # Crear el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"

        # Definir las columnas
        headers = [
            'CODIGO(SKU)', 'NOMBRE DEL PRODUCTO', 'STOCK', 'PRECIO', 'MARCA', 'ESTADO', 
            'MODELO', 'CATEGORIA', 'ANCHO'
        ]
        
        # Obtener todas las clases
        clases = Tblitemclase.objects.all()

        # Añadir los encabezados para las clases con su propiedad
        for clase in clases:
            headers.append(f"{clase.nombre} (Propiedad)")  # Añadimos columna con nombre de clase + propiedad

        # Escribir los encabezados
        ws.append(headers)

        # Obtener los productos
        productos = Tblitem.objects.all()

        # Recorrer los productos y agregar la información
        for producto in productos:
            row = [
                producto.codigosku,
                producto.titulo,
                producto.stock,
                producto.preciorebajado if producto.preciorebajado else producto.precionormal,
                producto.idmodelo.nombre if producto.idmodelo else "",
                producto.estado,
                producto.idmodelo.nombre if producto.idmodelo else "",
            ]

            # Obtener los IDs de las categorías asociadas al producto
            categorias_ids = producto.categoria_relacionada.values_list('idcategoria', flat=True)
            categoria_str = ", ".join(map(str, categorias_ids))  # Lista de IDs separados por coma
            row.append(categoria_str)  # Añadir la lista de IDs de categorías a la fila

            # Verificar si el producto está vinculado a alguna clase y agregar la propiedad
            for clase in clases:
                vinculo = tblitemclasevinculo.objects.filter(iditem=producto.pk, idclase=clase.idclase).first()
                if vinculo:
                    # Si está vinculado, colocar el valor de la propiedad en la celda
                    row.append(vinculo.propiedad)
                else:
                    row.append("")  # Si no está vinculado, dejar la celda vacía

            # Añadir la fila del producto
            ws.append(row)

        # Configurar los estilos
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        mandatory_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        optional_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        header_font = Font(bold=True)
        border_style = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Aplicar estilos a los encabezados
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar el ancho de las columnas
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in ws.iter_rows(min_row=2, max_row=len(productos) + 1, min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Preparar la respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_productos_vinculados.xlsx"'
        wb.save(response)
        return response
    
    
    
    @action(detail=False, methods=['get'], url_path='exportv2')
    def stream_excel_response(self, request):
        """
        Exportar productos optimizado con Pandas
        """

        # Obtener clases una sola vez
        clases = Tblitemclase.objects.values_list("idclase", "nombre")

        # Prefetch relaciones para evitar consultas múltiples
        productos = Tblitem.objects.select_related(
            "idmodelo"
        ).prefetch_related(
            Prefetch("categoria_relacionada"),
            Prefetch("clases_propiedades", queryset=tblitemclasevinculo.objects.select_related("idclase"))
).all()

        # Crear estructura de datos para Pandas
        data = []
        for producto in productos:
            row = {
                "CODIGO(SKU)": producto.codigosku,
                "NOMBRE DEL PRODUCTO": producto.titulo,
                "STOCK": producto.stock,
                "PRECIO": producto.preciorebajado if producto.preciorebajado else producto.precionormal,
                "MARCA": producto.idmodelo.nombre if producto.idmodelo else "",
                "ESTADO": producto.estado,
                "MODELO": producto.idmodelo.nombre if producto.idmodelo else "",
                "CATEGORIA": ", ".join(str(c.idcategoria.pk) for c in producto.categoria_relacionada.all()),
            }

            # Agregar propiedades de cada clase
            for idclase, nombre in clases:
                vinculo = next((v for v in producto.clases_propiedades.all() if v.idclase_id == idclase), None)
                row[f"{nombre}"] = vinculo.propiedad.upper() if vinculo else ""

            data.append(row)

        # Convertir datos en DataFrame
        df = pd.DataFrame(data)

        # Guardar en un buffer
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Plantilla", index=False)

            # Aplicar estilos con xlsxwriter
            workbook = writer.book
            worksheet = writer.sheets["Plantilla"]

            header_format = workbook.add_format(
                {"bold": True, "bg_color": "#FFC000", "border": 1, "align": "center"}
            )

            for col_num, value in enumerate(df.columns):
                worksheet.write(0, col_num, value, header_format)
                worksheet.set_column(col_num, col_num, 20)  # Ajustar ancho de columnas

        # Configurar respuesta HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_productos_vinculados.xlsx"'
        return response





class TblnoticiaViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Tblnoticia.objects.filter(activo=True).order_by('pk')
    serializer_class = TblnoticiaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['titulo','descripcion', 'estado']
    filterset_fields = ['activo', 'idnoticia', 'descripcion', 'estado', 'fechacreacion', 'fechamodificacion']


class TblsedeViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Tblsede.objects.filter(activo=True).order_by('pk')
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    serializer_class = TblsedeSerializer
    search_fields = ['direccion', 'nombre', 'email','departamento','provincia','distrito']
    filterset_fields = ['activo', 'direccion', 'nombre', 'email','departamento','provincia','distrito','telefono', 'direccion']


class TblpedidoViewSet(ModelViewSet):
    
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'sin_transaccion',
                openapi.IN_QUERY,
                description="Filtra los pedidos sin transacción asignada",
                type=openapi.TYPE_BOOLEAN
            )
        ]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    
    queryset = Tblpedido.objects.filter(activo=True).order_by('pk')
    serializer_class = TblpedidoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['idcliente__nombreusuario', 'total', 'estado']
    filterset_class = TblpedidoFilter
    def create(self, request, *args, **kwargs):
        """
        Sobrescribe el método create para procesar detalles del pedido.
        """
        data = request.data
        productos = data.pop('productos', None)  # Extraemos los productos del payload

        if not productos or not isinstance(productos, list):
            raise ValidationError({'productos': 'Debe incluir una lista de productos.'})
        
        with transaction.atomic():  # Usamos una transacción para garantizar consistencia
            pedido_serializer = self.get_serializer(data=data)
            pedido_serializer.is_valid(raise_exception=True)
            pedido = pedido_serializer.save()

            # Procesar los productos
            for producto in productos:
                idproducto = producto.get('idproduct')  # Obtener el ID del producto
                cantidad = producto.get('cantidad')
                print(idproducto)
                try:
                    producto_obj = Tblitem.objects.get(pk=idproducto)
                except Tblitem.DoesNotExist:
                    raise ValidationError({'error': f'El producto {idproducto} no existe.'})
                print(producto_obj.stock)
                # Validar stock
                if producto_obj.stock < cantidad:
                    raise ValidationError({'error': f'Stock insuficiente para {producto_obj.descripcion}.'})

                # Reducir stock del producto
                #producto_obj.stock -= cantidad
                #producto_obj.save()

                # Crear el detalle del pedido
                producto['idpedido'] = pedido.idpedido  # Asociamos el pedido
                detalle_serializer = TbldetallepedidoSerializer(data=producto)
                detalle_serializer.is_valid(raise_exception=True)
                detalle_serializer.save()

        return Response(pedido_serializer.data, status=status.HTTP_201_CREATED)
    
    
    @action(detail=True, methods=['post'], url_path='cancelar')
    def cancelar_pedido(self, request, pk=None):
        try:
            with transaction.atomic():
                pedido = self.get_object()

                if pedido.estado == 4:  # Suponiendo que 0 significa 'cancelado'
                    return Response({"error": "El pedido ya está cancelado."}, status=status.HTTP_400_BAD_REQUEST)

                # Obtener los detalles del pedido
                detalles = Tbldetallepedido.objects.filter(idpedido=pedido)
                print(detalles)
                for detalle in detalles:
                    producto = detalle.idproduct
                    producto.stock += detalle.cantidad
                    producto.save()

                # Cambiar estado del pedido a cancelado
                pedido.estado = 4
                pedido.save()

                return Response({"message": "Pedido cancelado y stock actualizado correctamente."}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TblCarruselViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = TblCarrusel.objects.filter(activo=True).order_by('pk')
    serializer_class = TblCarruselSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['titulo','descripcion', 'estado']
    filterset_fields = ['activo', 'id',  'descripcion', 'estado', 'fechacreacion', 'fechamodificacion']



class TblusuarioViewSet(ModelViewSet):
    queryset = CustomUser.objects.filter(activo=True).order_by('pk')
    serializer_class = CustomUserSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend, DateTimeIntervalFilter]
    search_fields = ['nombreusuario', 'nombre', 'apellidos', 'estado','email']
    filterset_fields = ['activo', 'nombreusuario', 'nombre', 'apellidos','departamento','provincia','distrito','telefono', 'estado', 'email_verified_at', 'direccion', 'fechacreacion', 'fechamodificacion','is_staff']



class TipocambioViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Tipocambio.objects.filter(activo=True).order_by('pk')
    serializer_class = TipocambioSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['tipocambio', 'idmoneda__nombre', 'fechacreacion', 'fechamodificacion']
    filterset_fields = ['activo', 'tipocambio', 'idmoneda__nombre', 'fechacreacion', 'fechamodificacion']




class ValoracionViewSet(ModelViewSet):
    queryset = Valoracion.objects.filter(activo=True).order_by('pk')
    serializer_class = ValoracionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['comentario', 'estrellas']
    filterset_fields = ['activo', 'idvaloracion', 'estrellas', 'comentario', 'estado', 'idproduct_id', 'fechacreacion', 'fechamodificacion']

    


class TbldetallecarritoViewSet(ModelViewSet):
    queryset = Tbldetallecarrito.objects.filter(activo=True).order_by('pk')
    serializer_class = TbldetallecarritoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['idproduct__descripcion', 'cantidad']
    filterset_fields = ['activo',  'idproduct_id', 'cantidad', 'isuser_id', 'idcupon_id']

class TbldetallecarritoViewSet(ModelViewSet):
    queryset = Tbldetallecarrito.objects.filter(activo=True).order_by('pk')
    serializer_class = TbldetallecarritoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['idproduct__descripcion', 'cantidad']
    filterset_fields = ['activo',  'idproduct_id', 'cantidad', 'isuser_id', 'idcupon_id']



class TblimagenitemViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Tblimagenitem.objects.filter(activo=True).order_by('pk')
    serializer_class = TblimagenitemSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['imagen', 'idproduct__descripcion']
    filterset_fields = ['activo', 'idimagen', 'idproduct_id', 'estado']
    
    @action(detail=False, methods=['post'], url_path='upload-multiple', serializer_class=MultipleImagenItemSerializer)
    def upload_multiple(self, request):
        """
        Endpoint personalizado para subir múltiples imágenes relacionadas a un producto.
        """
        serializer = MultipleImagenItemSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Imágenes subidas exitosamente'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




class MultipleImagenItemView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = MultipleImagenItemSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Imágenes subidas exitosamente'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    
class TblitemclaseViewSet(ModelViewSet):
    queryset = Tblitemclase.objects.filter(activo=True).order_by('pk')
    serializer_class = TblitemclaseSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre']
    filterset_fields = ['activo', 'idclase', 'nombre']

class tblitemcuponSerializerViewSet(ModelViewSet):
    queryset = tblitemcupon.objects.filter(activo=True).order_by('pk')
    serializer_class = tblitemcuponSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre']
    filterset_fields = ['activo', 'id', 'iditem_id', 'idcupon_id']



class tblitemclasevinculoViewSet(ModelViewSet):
    queryset = tblitemclasevinculo.objects.filter(activo=True).order_by('pk')
    serializer_class = TblitemclasevinculoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = [ 'propiedad', 'idclase__nombre']
    filterset_fields = ['activo', 'id', 'iditem_id', 'propiedad', 'idclase_id']



class TblitempropiedadViewSet(ModelViewSet):
    queryset = Tblitempropiedad.objects.filter(activo=True).order_by('pk')
    serializer_class = TblitempropiedadSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre', 'idclase__nombre']
    filterset_fields = ['activo', 'idpropiedad', 'nombre', 'idclase_id']
    @action(detail=False, methods=['get'], url_path='por-clase/(?P<clase_id>\d+)')
    def por_clase(self, request, clase_id=None):
        """
        Retorna todas las itempropiedades relacionadas a una clase de item específica.
        """
        propiedades = Tblitempropiedad.objects.filter(idclase=clase_id)
        serializer = TblitempropiedadSerializer(propiedades, many=True)
        return Response(serializer.data)




class TblitemrelacionadoViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = Tblitemrelacionado.objects.filter(activo=True).order_by('pk')
    serializer_class = TblitemrelacionadoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['item_id__descripcion']
    filterset_fields = ['activo', 'item_id', 'item_relacionado_id']


class tblitemcategoriaViewSet(ModelViewSet):
    permission_classes = [AllowAnyForReadOnly] 
    
    queryset = tblitemcategoria.objects.filter(activo=True).order_by('pk')
    serializer_class = tblitemcategoriaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['iditem_id__descripcion']
    filterset_fields = ['activo', 'iditem_id','idcategoria_id']




class TbldetallepedidoViewSet(ModelViewSet):
    queryset = Tbldetallepedido.objects.filter(activo=True).order_by('pk')
    serializer_class = TbldetallepedidoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['idpedido__id', 'idproduct__descripcion']
    filterset_fields = ['activo', 'idpedido_id', 'idproduct_id', 'cantidad', 'preciototal', 'preciunitario']
    

class TransaccionViewSet(ModelViewSet):
    queryset = tblTransaccion.objects.filter(activo=True).order_by('pk')
    serializer_class = TransaccionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['nombre_en_tarjeta']
    filterset_fields = [
    "metodo_pago",
    "nombre_en_tarjeta",
    "numero_tarjeta",
    "monto_total"
]
from rest_framework.permissions import AllowAny

User = get_user_model()

class AutoLoginView(APIView):
    permission_classes = [AllowAny]  # Permitir acceso sin autenticación

    def get(self, request):
        try:
            # Buscar al usuario con ID = 2
            user = User.objects.get(id=2)

            # Iniciar sesión con el usuario
            login(request, user)

            # Generar un nuevo token Knox para el usuario
            token_instance, token_key  = AuthToken.objects.create(user)  # Knox devuelve (instance, token)
            expiry = token_instance.expiry

            return Response({
                "success": True,
                "token": token_key  ,
                "expiry": expiry.isoformat()

                    })
        except User.DoesNotExist:
            return Response({
                "success": False,
                "error": {
                    "code": "user_not_found",
                    "message": "El usuario con ID 2 no existe."
                }
            }, status=404)

